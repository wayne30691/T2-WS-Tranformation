import streamlit as st
import pandas as pd
import re
import os, io

# ---------- Persist across reruns (optional) ----------
_PERSIST_PATH = "data/mapping.xlsx"

# preload saved mapping once per session
if "_mapping_init" not in st.session_state:
    if os.path.exists(_PERSIST_PATH):
        with open(_PERSIST_PATH, "rb") as f:
            st.session_state["_mapping_bytes"] = f.read()
            st.session_state["_mapping_name"] = "mapping.xlsx"
            st.session_state["_have_mapping"] = True
    st.session_state["_mapping_init"] = True

# sidebar controls
with st.sidebar:
    st.checkbox("Remember mapping between runs", value=True, key="_persist_mapping")
    if st.button("Clear saved mapping"):
        st.session_state.pop("_mapping_bytes", None)
        st.session_state.pop("_mapping_name", None)
        st.session_state["_have_mapping"] = False
        try:
            os.remove(_PERSIST_PATH)
        except FileNotFoundError:
            pass
        st.success("Cleared saved mapping.")

# ---------- Monkey patch ----------
_orig_file_uploader = st.file_uploader

class _MemoryUpload(io.BytesIO):
    def __init__(self, data: bytes, name="mapping.xlsx",
                 type_="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        super().__init__(data)
        self.name = name
        self.type = type_
        self.size = len(data)

def _file_uploader_with_memory(label, *args, **kwargs):
    key = kwargs.get("key", "")
    is_mapping = ("mapping" in str(label).lower()) or ("mapping" in str(key).lower())

    up = _orig_file_uploader(label, *args, **kwargs)

    # If this is the mapping uploader and a new file arrived: save + (optionally) persist
    if is_mapping and up is not None:
        data = up.read()
        st.session_state["_mapping_bytes"] = data
        st.session_state["_mapping_name"] = getattr(up, "name", "mapping.xlsx")
        st.session_state["_have_mapping"] = True
        if st.session_state.get("_persist_mapping", False):
            os.makedirs(os.path.dirname(_PERSIST_PATH), exist_ok=True)
            with open(_PERSIST_PATH, "wb") as f:
                f.write(data)
        return _MemoryUpload(st.session_state["_mapping_bytes"], name=st.session_state["_mapping_name"])

    # If this is the mapping uploader and nothing was uploaded this run, but we have it cached: return it
    if is_mapping and st.session_state.get("_have_mapping"):
        return _MemoryUpload(st.session_state["_mapping_bytes"], name=st.session_state.get("_mapping_name", "mapping.xlsx"))

    # Otherwise, behave like normal
    return up

st.file_uploader = _file_uploader_with_memory

# Streamlit app title
st.title("📊 T2 WS Transformations")
st.write("Upload an Excel file and choose the transformation format.")


# Select transformation format
transformation_choice = st.selectbox("Select Transformation Format:", ["30010008 利多吉", "30010010 酒倉盛豐行", "30010013 酒田", "30010017 正興(振興)", "30010031 廣茂隆(八條)", "30010059 誠邦有限公司", "30010061 向日葵", "30010085 宏酒樽 (夜)", "30010154 亨玖", 
                                                                       "30010176 振泰 ON", "30010185 瑞星翰德(夜點)", "30010199 振泰 OFF", "30010203 宏酒樽 (日)", "30010315 圳程", "30010316 大倉捷", "30020016 日嵩", "30020027 榮好(實儀)", "30020076 酒國英豪", 
                                                                       "30020145 鏵錡", "30020180 暐倫 OFF", "30020203 玄星 OFF", "30020216 久悅貿易", "30030021 合歡 ON", "30030061 合歡 OFF", "30030076 裕陞（分月）", "30030083 東瀛", "30030084 華恩", "30030088 九久", 
                                                                       "30030094 和易 ON", "30030106 明輝", "33001422 和易 OFF"])


if transformation_choice == "30010085 宏酒樽 (夜)":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="new_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="new_mapping")
    
    if raw_data_file is not None and mapping_file is not None:
        # Find the sheet that contains "夜" in the name
        xls = pd.ExcelFile(raw_data_file)
        sheet_name = next((sheet for sheet in xls.sheet_names if "夜" in sheet), None)

        if sheet_name:
            df_raw = xls.parse(sheet_name)
            
            sheets_mapping = pd.ExcelFile(mapping_file).sheet_names  
            dfs_mapping = {sheet: pd.read_excel(mapping_file, sheet_name=sheet) for sheet in sheets_mapping}
            
            df_transformed = df_raw.iloc[:, [1, 2, 3, 4, 5, 6]].copy()
            df_transformed.columns = ["Date", "Outlet Code", "Outlet Name", "Product Code", "Product Name", "Number of Bottles"]
            
            # Add fixed columns
            df_transformed.insert(0, "Column1", "INV")
            df_transformed.insert(1, "Column2", "U")
            df_transformed.insert(2, "Column3", "30010085")
            df_transformed.insert(3, "Column4", "宏酒樽 ON")
            
            df_transformed["Date"] = pd.to_datetime(df_transformed["Date"]).dt.strftime('%Y%m%d')
            
            # Map product codes
            df_sku_mapping = dfs_mapping["SKU Mapping"]
            df_sku_mapping = df_sku_mapping[["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")
            
            df_transformed = df_transformed.merge(
                df_sku_mapping,
                left_on="Product Code",
                right_on="ASI_CRM_Offtake_Product__c",
                how="left"
            )
            
            df_transformed.rename(columns={"ASI_CRM_SKU_Code__c": "SKU Code"}, inplace=True)
            df_transformed.drop(columns=["ASI_CRM_Offtake_Product__c"], inplace=True)
            
            # ✅ Fix Outlet Code Mapping Issue ✅
            df_transformed["Outlet Code"] = df_transformed["Outlet Code"].astype(str)

            # Optional replacement only if values are dates (skip if not needed)
            df_transformed["Outlet Code"] = df_transformed["Outlet Code"].replace({
                "2024-05-01 00:00:00": "5月1日",
                "2024-07-01 00:00:00": "7月1日",
                "2024-07-02 00:00:00": "07-02"
            })
            
            # ✅🔄 Updated Customer Mapping with 30010085 Filter
            df_customer_mapping = dfs_mapping["Customer Mapping"]
            df_customer_mapping = df_customer_mapping[
                df_customer_mapping["ASI_CRM_Mapping_Cust_No__c"] == 30010085
            ][["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"]].drop_duplicates(
                subset="ASI_CRM_Offtake_Customer_No__c"
            )
            
            df_transformed = df_transformed.merge(
                df_customer_mapping,
                left_on="Outlet Code",
                right_on="ASI_CRM_Offtake_Customer_No__c",
                how="left"
            )
            
            df_transformed.rename(columns={"ASI_CRM_JDE_Cust_No_Formula__c": "PRT Customer Code"}, inplace=True)
            df_transformed.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "Outlet Code"], inplace=True)
            
            # Reorder the columns
            column_order = ["Column1", "Column2", "Column3", "Column4", "PRT Customer Code", "Outlet Name", "Date", "SKU Code", "Product Code", "Product Name", "Number of Bottles"]
            df_transformed = df_transformed[column_order]

            # Preview data in Streamlit
            st.write("✅ Processed Data Preview:")
            st.dataframe(df_transformed)
            
            # Export without headers
            output_filename = "30010085 transformation.xlsx"
            df_transformed.to_excel(output_filename, index=False, header=False)
            
            with open(output_filename, "rb") as f:
                st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010203 宏酒樽 (日)":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="new_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="new_mapping")
    
    if raw_data_file is not None and mapping_file is not None:
        # Find the sheet that contains "日" in the name
        xls = pd.ExcelFile(raw_data_file)
        sheet_name = next((sheet for sheet in xls.sheet_names if "日" in sheet), None)

        if sheet_name:
            df_raw = xls.parse(sheet_name)
            
            sheets_mapping = pd.ExcelFile(mapping_file).sheet_names  
            dfs_mapping = {sheet: pd.read_excel(mapping_file, sheet_name=sheet) for sheet in sheets_mapping}
            
            df_transformed = df_raw.iloc[:, [1, 2, 3, 4, 5, 6]].copy()
            df_transformed.columns = ["Date", "Outlet Code", "Outlet Name", "Product Code", "Product Name", "Number of Bottles"]
            
            # Add fixed columns
            df_transformed.insert(0, "Column1", "INV")
            df_transformed.insert(1, "Column2", "U")
            df_transformed.insert(2, "Column3", "30010203")
            df_transformed.insert(3, "Column4", "宏酒樽 OFF")
            
            df_transformed["Date"] = pd.to_datetime(df_transformed["Date"]).dt.strftime('%Y%m%d')
            
            # Map product codes
            df_sku_mapping = dfs_mapping["SKU Mapping"]
            df_sku_mapping = df_sku_mapping[["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

            # Clean and normalize SKU columns
            df_transformed["Product Code"] = df_transformed["Product Code"].astype(str).str.strip().str.upper()
            df_sku_mapping["ASI_CRM_Offtake_Product__c"] = df_sku_mapping["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()

            df_transformed = df_transformed.merge(
                df_sku_mapping,
                left_on="Product Code",
                right_on="ASI_CRM_Offtake_Product__c",
                how="left"
            )
            
            df_transformed.rename(columns={"ASI_CRM_SKU_Code__c": "SKU Code"}, inplace=True)
            df_transformed.drop(columns=["ASI_CRM_Offtake_Product__c"], inplace=True)
            
            # ✅ Fix Outlet Code Mapping Issue ✅
            df_transformed["Outlet Code"] = df_transformed["Outlet Code"].astype(str)

            # Optional replacement only if values are dates (skip if not needed)
            df_transformed["Outlet Code"] = df_transformed["Outlet Code"].replace({
                "2024-05-01 00:00:00": "5月1日",
                "2024-07-01 00:00:00": "7月1日",
                "2024-07-02 00:00:00": "07-02"
            })
            
            # ✅🔄 Updated Customer Mapping with 30010085 Filter
            df_customer_mapping = dfs_mapping["Customer Mapping"]
            df_customer_mapping = df_customer_mapping[
                df_customer_mapping["ASI_CRM_Mapping_Cust_No__c"] == 30010203
            ][["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"]].drop_duplicates(
                subset="ASI_CRM_Offtake_Customer_No__c"
            )
            
            df_transformed = df_transformed.merge(
                df_customer_mapping,
                left_on="Outlet Code",
                right_on="ASI_CRM_Offtake_Customer_No__c",
                how="left"
            )
            
            df_transformed.rename(columns={"ASI_CRM_JDE_Cust_No_Formula__c": "PRT Customer Code"}, inplace=True)
            df_transformed.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "Outlet Code"], inplace=True)
            
            # Reorder the columns
            column_order = ["Column1", "Column2", "Column3", "Column4", "PRT Customer Code", "Outlet Name", "Date", "SKU Code", "Product Code", "Product Name", "Number of Bottles"]
            df_transformed = df_transformed[column_order]

            # Preview data in Streamlit
            st.write("✅ Processed Data Preview:")
            st.dataframe(df_transformed)
            
            # Export without headers
            output_filename = "30010203 transformation.xlsx"
            df_transformed.to_excel(output_filename, index=False, header=False)
            
            with open(output_filename, "rb") as f:
                st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010061 向日葵":
    uploaded_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="sunflower_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="sunflower_mapping")

    if uploaded_file is not None and mapping_file is not None:
        df = pd.read_excel(uploaded_file, header=None)

        # Create an empty list to store the extracted data
        data = []

        # Initialize variables to hold the current customer name, code, and date
        current_customer = None
        current_customer_code = None
        current_date = None

        # Start processing from row 8 (index 7)
        for i in range(7, len(df)):
            row = df.iloc[i]

            if isinstance(row[0], str) and '客戶名稱' in row[0]:
                cleaned_text = re.sub(r'[\u200b\ufeff]', '', row[0]).strip()
                match = re.search(r'客戶編號[:：]\s*([\d\-]+).*客戶名稱[:：]\s*(.*)', cleaned_text)
                if match:
                    current_customer_code = match.group(1).strip()
                    current_customer = match.group(2).strip()

            if isinstance(row[0], str) and re.match(r'\d{3}/\d{2}/\d{2}', row[0]):
                year, month, day = map(int, row[0].split('/'))
                current_date = f'{year + 1911}{month:02}{day:02}'

            if pd.notna(row[1]):
                product_code = row[1]
                product_name = row[2]
                quantity = row[3]

                data.append([current_customer_code, current_customer, current_date, product_code, product_name, quantity])

        result_df = pd.DataFrame(data, columns=['Customer Code', 'Customer Name', 'Date', 'Product Code', 'Product Name', 'Quantity'])

        # Add fixed columns
        result_df.insert(0, 'Column1', 'INV')
        result_df.insert(1, 'Column2', 'U')
        result_df.insert(2, 'Column3', '30010061')
        result_df.insert(3, 'Column4', '向日葵')

        # --- ✅ CUSTOMER MAPPING ---
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        df_customer = dfs_mapping["Customer Mapping"]
        df_customer = df_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        result_df = result_df.merge(
            df_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        result_df["Customer Code"] = result_df["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        result_df.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # --- ✅ SKU MAPPING ---
        df_sku = dfs_mapping["SKU Mapping"]
        df_sku = df_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        result_df = result_df.merge(
            df_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = result_df.columns.get_loc("Product Code")
        result_df.insert(product_index, "PRT Product Code", result_df["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        result_df.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Preview data in Streamlit
        st.write("✅ Processed Data Preview:")
        st.dataframe(result_df)

        output_filename = "30010061 transformation.xlsx"
        result_df.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010010 酒倉盛豐行":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="sakakura_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="sakakura_mapping")

    if raw_data_file and mapping_file:
        raw_df = pd.read_excel(raw_data_file, sheet_name=0, header=None)
        # Extract date from cell A5
        date_string = str(raw_df.iloc[4, 0])
        match = re.search(r'至\s*(\d{3}/\d{2}/\d{2})', date_string)
        if match:
            roc_date = match.group(1)
            year, month, day = map(int, roc_date.split('/'))
            final_date = f"{year + 1911}{month:02d}{day:02d}"
        else:
            final_date = None

        current_product_code = None
        current_product_name = None
        data = []

        for _, row in raw_df.iterrows():
            col_a = str(row[0]).strip() if pd.notna(row[0]) else ""
            col_b = str(row[1]).strip() if pd.notna(row[1]) else ""
            col_d = row[3] if pd.notna(row[3]) else None

            if "貨品編號" in col_a and "貨品名稱" in col_a:
                match = re.search(r'貨品編號[:：]([A-Z0-9\-]+)\s+貨品名稱[:：](.+)', col_a)
                if match:
                    current_product_code = match.group(1).strip()
                    current_product_name = match.group(2).strip()
                continue

            if "小計" in col_a or "小計" in col_b:
                continue

            if col_a and col_b and isinstance(col_d, (int, float)) and current_product_code:
                data.append([
                    col_a, col_b, final_date,
                    current_product_code, current_product_name,
                    int(col_d)
                ])

        df_cleaned = pd.DataFrame(data, columns=[
            "Customer Code", "Customer Name", "Date",
            "Product Code", "Product Name", "Quantity"
        ])

        mapping_customer = pd.read_excel(mapping_file, sheet_name="Customer Mapping")
        mapping_customer = mapping_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_cleaned = df_cleaned.merge(
            mapping_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_cleaned["Customer Code"] = df_cleaned["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        df_cleaned.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        mapping_sku = pd.read_excel(mapping_file, sheet_name="SKU Mapping")
        mapping_sku = mapping_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_cleaned = df_cleaned.merge(
            mapping_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_code_index = df_cleaned.columns.get_loc("Product Code")
        df_cleaned.insert(product_code_index, "PRT Product Code", df_cleaned["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_cleaned.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        df_cleaned.insert(0, "Column1", "INV")
        df_cleaned.insert(1, "Column2", "U")
        df_cleaned.insert(2, "Column3", "30010010")
        df_cleaned.insert(3, "Column4", "酒倉 ON")

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_cleaned)

        output_filename = "30010010 transformation.xlsx"
        df_cleaned.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010013 酒田":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx", "xls"], key="sakata_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="sakata_mapping")

    if raw_data_file and mapping_file:
        raw_df = pd.read_excel(raw_data_file, sheet_name=0, header=None)  # Use first sheet

        # Extract ROC date from cell A5
        date_string = str(raw_df.iloc[4, 0])
        match = re.search(r'至\s*(\d{3}/\d{2}/\d{2})', date_string)
        if match:
            roc_date = match.group(1)
            year, month, day = map(int, roc_date.split('/'))
            final_date = f"{year + 1911}{month:02d}{day:02d}"
        else:
            final_date = None

        current_product_code = None
        current_product_name = None
        data = []

        for _, row in raw_df.iterrows():
            col_a = str(row[0]).strip() if pd.notna(row[0]) else ""
            col_b = str(row[1]).strip() if pd.notna(row[1]) else ""
            col_f = row.iloc[5] if len(row) > 5 and pd.notna(row.iloc[5]) else None  # SAFE

            if "貨品編號" in col_a and "貨品名稱" in col_a:
                match = re.search(r'貨品編號[:：]([A-Z0-9\-]+)\s+貨品名稱[:：](.+)', col_a)
                if match:
                    current_product_code = match.group(1).strip()
                    current_product_name = match.group(2).strip()
                continue

            if "小計" in col_a or "小計" in col_b:
                continue

            if re.match(r'^[A-Z]', col_a):  # Allow any valid Latin-starting code
                if col_f and isinstance(col_f, (int, float)):
                    data.append([
                        col_a, col_b, final_date,
                        current_product_code, current_product_name,
                        int(col_f)
                    ])

        df_cleaned = pd.DataFrame(data, columns=[
            "Customer Code", "Customer Name", "Date",
            "Product Code", "Product Name", "Quantity"
        ])

        # Load customer mapping
        mapping_customer = pd.read_excel(mapping_file, sheet_name="Customer Mapping")
        mapping_customer = mapping_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_cleaned = df_cleaned.merge(
            mapping_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_cleaned["Customer Code"] = df_cleaned["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        df_cleaned.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # Load SKU mapping
        mapping_sku = pd.read_excel(mapping_file, sheet_name="SKU Mapping")
        mapping_sku = mapping_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_cleaned = df_cleaned.merge(
            mapping_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_code_index = df_cleaned.columns.get_loc("Product Code")
        df_cleaned.insert(product_code_index, "PRT Product Code", df_cleaned["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_cleaned.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Insert fixed identifier columns
        df_cleaned.insert(0, "Column1", "INV")
        df_cleaned.insert(1, "Column2", "U")
        df_cleaned.insert(2, "Column3", "30010013")
        df_cleaned.insert(3, "Column4", "酒田 ON")

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_cleaned)

        output_filename = "30010013 transformation.xlsx"
        df_cleaned.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010059 誠邦有限公司":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="raw_30010059")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="mapping_30010059")

    if raw_data_file is not None and mapping_file is not None:
        import re
        import pandas as pd

        # --- helper: remove trailing .0 and invisible spaces ---
        def clean_code(series: pd.Series) -> pd.Series:
            return (
                series.astype(str)
                      .str.replace(r"[\u00A0\u2007\u202F\u3000]", "", regex=True)  # NBSP/thin/full-width spaces
                      .str.strip()
                      .str.replace(r"\.0+$", "", regex=True)  # drop ONLY trailing .0/.00...
            )

        # ---------- read raw ----------
        raw_df = pd.read_excel(raw_data_file, sheet_name=0, header=None)

        # Step 1: detect format A/B
        offset = 0
        for i in range(10, len(raw_df)):
            row = raw_df.iloc[i]
            col_a = str(row[0]).strip() if pd.notna(row[0]) else ""
            col_b = str(row[1]).strip() if pd.notna(row[1]) else ""
            if re.match(r"\d{4}/\d{2}/\d{2}|\d{3}/\d{2}/\d{2}", col_a):
                offset = 0 if col_b.startswith("\u92b7") else 1  # 銷
                break

        # Step 2: extract transactions
        data = []
        current_product_code = None
        current_product_name = None
        found_first_product = False

        for i in range(len(raw_df)):
            row = raw_df.iloc[i]
            col_a = str(row[0]).strip() if pd.notna(row[0]) else ""
            col_b = str(row[1 - offset]).strip() if pd.notna(row[1 - offset]) else ""
            col_c = str(row[2 - offset]).strip() if pd.notna(row[2 - offset]) else ""
            col_d = str(row[3 - offset]).strip() if pd.notna(row[3 - offset]) else ""
            col_e = row[4 - offset] if pd.notna(row[4 - offset]) else None

            col_a_clean = col_a.replace('\u3000', ' ').replace('\xa0', ' ').strip()

            # Match both 【】 and []
            if "貨品編號:" in col_a_clean:
                m = re.search(r"貨品編號:\s*[\[\【]([^\]\】]+)[\]\】]\s*(.+)", col_a_clean)
                if m:
                    current_product_code = m.group(1).strip()
                    current_product_name = m.group(2).strip()
                    found_first_product = True
                continue

            if not found_first_product:
                continue

            if ("合計" in col_a_clean) or ("小計" in col_a_clean):
                continue

            if col_c and isinstance(col_e, (int, float)) and current_product_code and current_product_name:
                try:
                    y, mth, d = map(int, col_a_clean.split("/"))
                    if y < 1911:
                        y += 1911
                    gregorian_date = f"{y}{mth:02d}{d:02d}"
                except Exception:
                    gregorian_date = col_a_clean

                data.append([
                    col_c, col_d, gregorian_date,
                    current_product_code, current_product_name,
                    int(col_e)
                ])

        df_cleaned = pd.DataFrame(data, columns=[
            "Customer Code", "Customer Name", "Date",
            "Product Code", "Product Name", "Quantity"
        ])

        # ---------- load mappings ----------
        xlsx = pd.ExcelFile(mapping_file)
        dfs_mapping = {sheet: pd.read_excel(mapping_file, sheet_name=sheet) for sheet in xlsx.sheet_names}

        # Customer mapping: MERGE -> REWRITE Customer Code -> DROP extras
        df_customer = dfs_mapping["Customer Mapping"][[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_cleaned = df_cleaned.merge(
            df_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_cleaned["Customer Code"] = clean_code(df_cleaned["ASI_CRM_JDE_Cust_No_Formula__c"])
        df_cleaned.drop(
            columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"],
            inplace=True
        )

        # SKU mapping
        df_sku_mapping = dfs_mapping["SKU Mapping"][[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_cleaned = df_cleaned.merge(
            df_sku_mapping,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = df_cleaned.columns.get_loc("Product Code")
        df_cleaned.insert(
            product_index,
            "PRT Product Code",
            clean_code(df_cleaned["ASI_CRM_SKU_Code__c"])
        )
        df_cleaned.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # ---------- fixed columns + final ----------
        fixed_df = pd.DataFrame({
            "Column1": ["INV"] * len(df_cleaned),
            "Column2": ["U"] * len(df_cleaned),
            "Column3": ["30010059"] * len(df_cleaned),
            "Column4": ["誠邦有限公司"] * len(df_cleaned)
        })

        df_final = pd.concat([fixed_df, df_cleaned], axis=1)

        # Final safety: ensure Customer Code is text and clean
        df_final["Customer Code"] = clean_code(df_final["Customer Code"]).astype("string")

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "processed_30010059.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📅 Download Processed File", data=f, file_name=output_filename)


elif transformation_choice == "30010315 圳程":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="zc_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="zc_mapping")

    if raw_data_file and mapping_file:
        import openpyxl

        wb = openpyxl.load_workbook(raw_data_file, data_only=True)
        ws = wb.active

        # Try B3, then B4 if B3 is empty
        report_date_raw = ""
        for cell in ["B3", "B4"]:
            val = ws[cell].value
            if val:
                report_date_raw = str(val).strip()
                break

        # Parse the date string if available
        report_date = ""
        if "~" in report_date_raw:
            right_date = report_date_raw.split("~")[-1].strip()
            if len(right_date.split("/")) == 3:
                y, m, d = right_date.split("/")
                report_date = f"{int(y):04}{int(m):02}{int(d):02}"



        records = []
        product_name = product_code = customer_name = customer_code = None

        for i in range(ws.max_row):
            b = str(ws.cell(row=i+1, column=2).value).strip() if ws.cell(row=i+1, column=2).value else ""
            c = str(ws.cell(row=i+1, column=3).value).strip() if ws.cell(row=i+1, column=3).value else ""
            e = ws.cell(row=i+1, column=5).value if ws.cell(row=i+1, column=5).value else None

            if "(" in b and ")" in b:
                last_open = b.rfind("(")
                last_close = b.rfind(")")
                code = b[last_open + 1 : last_close]
                name = b[:last_open].strip()

                if i+2 < ws.max_row and str(ws.cell(row=i+2, column=2).value).strip() == "單據類別":
                    customer_name = name
                    customer_code = code
                else:
                    product_name = name
                    product_code = code

            if b == "出貨單" and c and isinstance(e, (int, float)):
                records.append({
                    "Customer Code": customer_code,
                    "Customer Name": customer_name,
                    "Date": report_date,
                    "Product Code": product_code,
                    "Product Name": product_name,
                    "Quantity": int(e),
                    "Document Number": c
                })

        df_transformed = pd.DataFrame(records)
        df_transformed.insert(0, "Column1", "INV")
        df_transformed.insert(1, "Column2", "U")
        df_transformed.insert(2, "Column3", "30010315")
        df_transformed.insert(3, "Column4", "圳程有限公司")

        # Load mappings
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        # Customer mapping
        df_customer = dfs_mapping["Customer Mapping"]
        df_customer = df_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_transformed = df_transformed.merge(
            df_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_transformed["Customer Code"] = df_transformed["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df_transformed.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # SKU mapping
        df_sku = dfs_mapping["SKU Mapping"]
        df_sku = df_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_transformed = df_transformed.merge(
            df_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = df_transformed.columns.get_loc("Product Code")
        df_transformed.insert(product_index, "PRT Product Code", df_transformed["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_transformed.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Reorder for consistency
        column_order = ["Column1", "Column2", "Column3", "Column4", "Customer Code", "Customer Name", "Date", "PRT Product Code", "Product Code", "Product Name", "Quantity", "Document Number"]
        df_transformed = df_transformed[column_order]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_transformed)

        output_filename = "30010315_transformation.xlsx"
        df_transformed.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)
            
elif transformation_choice == "30030088 九久":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="jj_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="jj_mapping")

    if raw_data_file and mapping_file:
        import openpyxl

        df_raw = pd.read_excel(raw_data_file, sheet_name=0, header=None)
        extracted_data = []

        i = 0
        while i < len(df_raw):
            row = df_raw.iloc[i, 0]

            if isinstance(row, str) and row.startswith("貨品編號:"):
                product_code = row.replace("貨品編號:", "").split()[0].strip()
                product_name = row.replace("貨品編號:", "").split(maxsplit=1)[1].strip() if len(row.split()) > 1 else ""

                data_start = i + 5
                while data_start < len(df_raw):
                    entry = df_raw.iloc[data_start]

                    if isinstance(entry[0], str) and entry[0].startswith("貨品編號:"):
                        break

                    # ✅ Skip if inbound: check if column E is '進貨單'
                    if str(entry[4]).strip() == "進貨單":
                        data_start += 1
                        continue

                    # Check if entry is valid (i.e., not empty)
                    if pd.isna(entry[0]) or pd.isna(entry[1]) or pd.isna(entry[2]):
                        data_start += 1
                        continue

                    # Initialize return flag
                    is_return = False

                    # ✅ Skip if inbound: check if column E is '進貨單'
                    # ✅ If it's a return '銷退單', we mark it and negate quantity later
                    doc_type = str(entry[4]).strip()
                    if doc_type == "進貨單":
                        data_start += 1
                        continue
                    elif doc_type == "銷退單":
                        is_return = True

                    try:
                        report_date = entry[0]
                        document_number = entry[1]
                        customer_code = entry[2]
                        customer_name = entry[3]
                        quantity = entry[6]

                        if pd.notna(quantity) and isinstance(quantity, (int, float)):
                            if is_return:
                                quantity = -abs(int(quantity))  # Ensure it's negative
                            extracted_data.append({
                                "Customer Code": str(customer_code).strip().split(".")[0],
                                "Customer Name": str(customer_name).strip(),
                                "Date": report_date,
                                "Product Code": product_code,
                                "Product Name": product_name,
                                "Quantity": int(quantity),
                                "Document Number": document_number
                            })
                    except Exception:
                        pass
                    data_start += 1
            i += 1


        df_transformed = pd.DataFrame(extracted_data)

        # Convert Minguo date to Gregorian YYYYMMDD
        def convert_minguo_date(minguo_str):
            try:
                parts = str(minguo_str).split('/')
                if len(parts) != 3:
                    return None
                year = int(parts[0]) + 1911
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}{month:02d}{day:02d}"
            except Exception:
                return None

        df_transformed["Date"] = df_transformed["Date"].apply(convert_minguo_date)

        # Add fixed columns
        df_transformed.insert(0, "Column4", "九久")
        df_transformed.insert(0, "Column3", "30030088")
        df_transformed.insert(0, "Column2", "U")
        df_transformed.insert(0, "Column1", "INV")

        # Load mapping sheets
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        # Customer mapping
        df_customer = dfs_mapping["Customer Mapping"]
        df_customer = df_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_transformed = df_transformed.merge(
            df_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_transformed["Customer Code"] = df_transformed["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df_transformed.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # SKU mapping
        df_sku = dfs_mapping["SKU Mapping"]
        df_sku = df_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_transformed = df_transformed.merge(
            df_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = df_transformed.columns.get_loc("Product Code")
        df_transformed.insert(product_index, "PRT Product Code", df_transformed["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_transformed.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Final column order
        column_order = ["Column1", "Column2", "Column3", "Column4", "Customer Code", "Customer Name", "Date", "PRT Product Code", "Product Code", "Product Name", "Quantity", "Document Number"]
        df_transformed = df_transformed[column_order]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_transformed)

        output_filename = "30030088_transformation.xlsx"
        df_transformed.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)


elif transformation_choice == "30020145 鏵錡":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xlsx"], key="30020145_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="30020145_mapping")

    if raw_data_file and mapping_file:
        import pandas as pd
        import re

        def extract_product_data_from_workbook(file):
            xls = pd.ExcelFile(file)
            combined_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)

                merged_cell_value = str(df.iloc[2, 0])
                product_match = re.search(r"貨品編號[:：]([A-Z0-9\-]+)\s+(.*)", merged_cell_value)

                if not product_match:
                    continue

                product_code = product_match.group(1).strip()
                product_name = product_match.group(2).strip()

                df_data = df.iloc[8:, :8].copy()
                df_data.columns = ["Date", "Document No", "Customer Code", "Distributor", "Customer Name", "Quantity", "Unit", "Note"]

                for _, row in df_data.iterrows():
                    if pd.isna(row["Date"]) or pd.isna(row["Customer Code"]) or pd.isna(row["Quantity"]):
                        continue

                    combined_data.append({
                        "Customer Code": row["Customer Code"],
                        "Customer Name": row["Customer Name"],
                        "Date": row["Date"],
                        "Product Code": product_code,
                        "Product Name": product_name,
                        "Quantity": row["Quantity"],
                        "Document No": row["Document No"]
                    })

            return pd.DataFrame(combined_data)

        def convert_minguo_to_gregorian(date_str):
            try:
                parts = str(date_str).split("/")
                if len(parts) != 3:
                    return None
                year = int(parts[0]) + 1911
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}{month:02d}{day:02d}"
            except:
                return None

        df_combined = extract_product_data_from_workbook(raw_data_file)
        df_combined["Date"] = df_combined["Date"].apply(convert_minguo_to_gregorian)

        # Load mapping sheets
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        # Customer Mapping
        df_customer = dfs_mapping["Customer Mapping"]
        df_customer = df_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_combined = df_combined.merge(
            df_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_combined["Customer Code"] = df_combined["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df_combined.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # SKU Mapping
        df_sku = dfs_mapping["SKU Mapping"]
        df_sku = df_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_combined = df_combined.merge(
            df_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = df_combined.columns.get_loc("Product Code")
        df_combined.insert(product_index, "PRT Product Code", df_combined["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_combined.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Insert fixed columns
        df_combined.insert(0, "Column4", "任我行")
        df_combined.insert(0, "Column3", "30020145")
        df_combined.insert(0, "Column2", "U")
        df_combined.insert(0, "Column1", "INV")

        # Preview result
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_combined)

        output_filename = "30020145_transformation.xlsx"
        df_combined.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010199 振泰 OFF":
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xls"], key="zhen_tai_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="zhen_tai_mapping")

    if raw_data_file is not None and mapping_file is not None:
        def extract_from_date_sheets(file):
            xls = pd.ExcelFile(file)
            all_data = []
            sheet_dates = {}

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)
                product_code = None
                product_name = None

                # ✅ Skip sheet if A5 is missing
                if df.shape[0] <= 4 or pd.isna(df.iloc[4, 0]):
                    continue

                # Extract date from A5
                raw_date_cell = str(df.iloc[4, 0])

                if "至" in raw_date_cell:
                    raw_date = raw_date_cell.split("至")[1].strip()
                    try:
                        parts = raw_date.split("/")
                        year = int(parts[0]) + 1911
                        month = int(parts[1])
                        day = int(parts[2])
                        formatted_date = f"{year:04d}{month:02d}{day:02d}"
                    except:
                        formatted_date = None
                else:
                    formatted_date = None
                sheet_dates[sheet_name] = formatted_date

                for i in range(len(df)):
                    cell_value = str(df.iloc[i, 0]).strip()
                    if cell_value.startswith("貨品編號:"):
                        rest = cell_value.replace("貨品編號:", "", 1).strip()
                        parts = rest.split("貨品名稱:")
                        product_code = parts[0].strip()
                        product_name = parts[1].strip() if len(parts) > 1 else ""
                        continue
                    if "小計" in cell_value or product_code is None:
                        continue

                    customer_code = str(df.iloc[i, 0]).strip()
                    customer_name = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
                    quantity = df.iloc[i, 2] if pd.notna(df.iloc[i, 2]) else None

                    if customer_code and quantity and isinstance(quantity, (int, float)):
                        all_data.append({
                            "Sheet": sheet_name,
                            "Customer Code": customer_code,
                            "Customer Name": customer_name,
                            "Date": formatted_date,
                            "Product Code": product_code,
                            "Product Name": product_name,
                            "Quantity": quantity
                        })

            return pd.DataFrame(all_data)

        df = extract_from_date_sheets(raw_data_file)

        # Mapping setup
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        # Filter customer mapping
        df_customer_mapping = dfs_mapping["Customer Mapping"]
        df_customer_mapping = df_customer_mapping[
            df_customer_mapping["ASI_CRM_Mapping_Cust_No__c"] == 30010199
        ][["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"]].drop_duplicates()

        df = df.merge(
            df_customer_mapping,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df["Customer Code"] = df["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        df_sku_mapping = dfs_mapping["SKU Mapping"]
        df_sku_mapping = df_sku_mapping[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates()

        df = df.merge(
            df_sku_mapping,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )
        df.insert(df.columns.get_loc("Product Code"), "PRT Product Code", df["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Add 4 fixed columns
        df.insert(1, "Col1", "INV")
        df.insert(2, "Col2", "U")
        df.insert(3, "Col3", "30010199")
        df.insert(4, "Col4", "振泰 OFF")

        # Optional: Toggle by Month (📅 grouped by available months)
        available_months = sorted(set([d[:6] for d in df["Date"].dropna().astype(str)]))
        month_filter = st.radio("📅 Filter by Month:", ["All"] + available_months)

        if month_filter != "All":
            df = df[df["Date"].astype(str).str.startswith(month_filter)]

        # Final column order (excluding 'Sheet')
        df = df[[
            "Col1", "Col2", "Col3", "Col4",
            "Customer Code", "Customer Name", "Date",
            "PRT Product Code", "Product Code", "Product Name", "Quantity"
        ]]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df)

        # Export to Excel (remove first row, no headers)
        output_filename = "30010199_transformation.xlsx"
        df_export = df.copy()
        df_export.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )

elif transformation_choice == "30010176 振泰 ON":
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xls"], key="zhen_tai_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xlsx"], key="zhen_tai_mapping")

    if raw_data_file is not None and mapping_file is not None:
        def extract_from_date_sheets(file):
            xls = pd.ExcelFile(file)
            all_data = []
            sheet_dates = {}

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)
                product_code = None
                product_name = None

                # ✅ Skip sheet if A5 is missing
                if df.shape[0] <= 4 or pd.isna(df.iloc[4, 0]):
                    continue

                # Extract date from A5
                raw_date_cell = str(df.iloc[4, 0])

                if "至" in raw_date_cell:
                    raw_date = raw_date_cell.split("至")[1].strip()
                    try:
                        parts = raw_date.split("/")
                        year = int(parts[0]) + 1911
                        month = int(parts[1])
                        day = int(parts[2])
                        formatted_date = f"{year:04d}{month:02d}{day:02d}"
                    except:
                        formatted_date = None
                else:
                    formatted_date = None
                sheet_dates[sheet_name] = formatted_date

                for i in range(len(df)):
                    cell_value = str(df.iloc[i, 0]).strip()
                    if cell_value.startswith("貨品編號:"):
                        rest = cell_value.replace("貨品編號:", "", 1).strip()
                        parts = rest.split("貨品名稱:")
                        product_code = parts[0].strip()
                        product_name = parts[1].strip() if len(parts) > 1 else ""
                        continue
                    if "小計" in cell_value or product_code is None:
                        continue

                    customer_code = str(df.iloc[i, 0]).strip()
                    customer_name = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
                    quantity = df.iloc[i, 2] if pd.notna(df.iloc[i, 2]) else None

                    if customer_code and quantity and isinstance(quantity, (int, float)):
                        all_data.append({
                            "Sheet": sheet_name,
                            "Customer Code": customer_code,
                            "Customer Name": customer_name,
                            "Date": formatted_date,
                            "Product Code": product_code,
                            "Product Name": product_name,
                            "Quantity": quantity
                        })

            return pd.DataFrame(all_data)

        df = extract_from_date_sheets(raw_data_file)

        # Mapping setup
        dfs_mapping = {
            sheet: pd.read_excel(mapping_file, sheet_name=sheet)
            for sheet in pd.ExcelFile(mapping_file).sheet_names
        }

        # Filter customer mapping
        df_customer_mapping = dfs_mapping["Customer Mapping"]
        df_customer_mapping = df_customer_mapping[
            df_customer_mapping["ASI_CRM_Mapping_Cust_No__c"] == 30010176
        ][["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"]].drop_duplicates()

        df = df.merge(
            df_customer_mapping,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df["Customer Code"] = df["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        df_sku_mapping = dfs_mapping["SKU Mapping"]
        df_sku_mapping = df_sku_mapping[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates()

        df = df.merge(
            df_sku_mapping,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )
        df.insert(df.columns.get_loc("Product Code"), "PRT Product Code", df["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Add 4 fixed columns
        df.insert(1, "Col1", "INV")
        df.insert(2, "Col2", "U")
        df.insert(3, "Col3", "30010176")
        df.insert(4, "Col4", "振泰 ON")

        # Optional: Toggle by Month (📅 grouped by available months)
        available_months = sorted(set([d[:6] for d in df["Date"].dropna().astype(str)]))
        month_filter = st.radio("📅 Filter by Month:", ["All"] + available_months)

        if month_filter != "All":
            df = df[df["Date"].astype(str).str.startswith(month_filter)]

        # Final column order without 'Sheet'
        df = df[[
            "Col1", "Col2", "Col3", "Col4",
            "Customer Code", "Customer Name", "Date",
            "PRT Product Code", "Product Code", "Product Name", "Quantity"
        ]]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df)

        # Export to Excel (remove first row, no headers)
        output_filename = "30010176_transformation.xlsx"
        df_export = df.copy()
        df_export.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )

elif transformation_choice == "30030094 和易 ON":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xls", "xlsx"], key="heyi_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xls", "xlsx"], key="heyi_mapping")

    if raw_data_file and mapping_file:
        raw_df = pd.read_excel(raw_data_file, sheet_name="Page 1", header=None)

        # Extract depletion rows with context
        extracted_data = []
        product_code = None
        product_name = None

        for idx, row in raw_df.iterrows():
            col0 = str(row[0]) if pd.notna(row[0]) else ""
            col3 = str(row[3]) if pd.notna(row[3]) else ""

            if col0.startswith("產品編號:"):
                product_code = col0.replace("產品編號:", "").strip()

            if col3.startswith("品名規格:"):
                product_name = col3.replace("品名規格:", "").strip()

            if str(row[3]).strip() == "銷貨（庫存）":
                report_date = row[0]
                document_number = row[1]
                customer_name = row[2]
                quantity = row[5]
                customer_code = row[9]

                if all(pd.notna([report_date, document_number, customer_name, quantity, customer_code])):
                    extracted_data.append({
                        "Customer Code": str(customer_code).strip(),
                        "Customer Name": str(customer_name).strip(),
                        "Date": report_date,
                        "Product Code": product_code,
                        "Product Name": product_name,
                        "Quantity": int(quantity),
                        "Document Number": document_number
                    })

        depletion_df = pd.DataFrame(extracted_data)

        # Add fixed columns
        depletion_df.insert(0, "INV", "INV")
        depletion_df.insert(1, "U", "U")
        depletion_df.insert(2, "Customer Group Code", "30030094")
        depletion_df.insert(3, "Customer Group Name", "和易 ON")

        # Mapping: Customer
        mapping_customer = pd.read_excel(mapping_file, sheet_name="Customer Mapping")
        mapping_customer = mapping_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        depletion_df = depletion_df.merge(
            mapping_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        depletion_df["Customer Code"] = depletion_df["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\\.0$", "", regex=True)
        depletion_df.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # Mapping: SKU
        mapping_sku = pd.read_excel(mapping_file, sheet_name="SKU Mapping")
        mapping_sku = mapping_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        depletion_df = depletion_df.merge(
            mapping_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = depletion_df.columns.get_loc("Product Code")
        depletion_df.insert(product_index, "PRT Product Code", depletion_df["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        depletion_df.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        # Convert Minguo date to YYYYMMDD
        def convert_minguo_date(date_str):
            try:
                if isinstance(date_str, str) and '/' in date_str:
                    parts = date_str.strip().split('/')
                    year = int(parts[0]) + 1911
                    month = int(parts[1])
                    day = int(parts[2])
                    return f"{year:04d}{month:02d}{day:02d}"
                return date_str
            except:
                return date_str

        depletion_df["Date"] = depletion_df["Date"].apply(convert_minguo_date)

        st.write("✅ Processed Data Preview:")
        st.dataframe(depletion_df)

        output_filename = "30030094_transformation.xlsx"
        depletion_df.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "33001422 和易 OFF":
    raw_data_file = st.file_uploader("Upload Raw Sales Data", type=["xls", "xlsx"], key="heyi_off_raw")
    mapping_file = st.file_uploader("Upload Mapping File", type=["xls", "xlsx"], key="heyi_off_mapping")

    if raw_data_file and mapping_file:
        raw_df = pd.read_excel(raw_data_file, sheet_name="Page 1", header=None)

        extracted_data = []
        product_code = None
        product_name = None

        for _, row in raw_df.iterrows():
            col0 = str(row[0]) if pd.notna(row[0]) else ""
            col3 = str(row[3]) if pd.notna(row[3]) else ""

            if col0.startswith("產品編號:"):
                product_code = col0.replace("產品編號:", "").strip()

            if col3.startswith("品名規格:"):
                product_name = col3.replace("品名規格:", "").strip()

            if str(row[3]).strip() in ["銷貨（庫存）", "銷貨退回"]:
                report_date = row[0]
                document_number = row[1]
                customer_name = row[2]
                quantity = row[5]
                customer_code = row[9]

                if all(pd.notna([report_date, document_number, customer_name, quantity, customer_code])):
                    qty = int(quantity)
                    if str(row[3]).strip() == "銷貨退回":
                        qty = -qty
                    extracted_data.append({
                        "Customer Code": str(customer_code).strip(),
                        "Customer Name": str(customer_name).strip(),
                        "Date": report_date,
                        "Product Code": product_code,
                        "Product Name": product_name,
                        "Quantity": qty,
                        "Document Number": document_number
                    })

        df_extracted = pd.DataFrame(extracted_data)

        # Add 4 fixed metadata columns
        df_extracted.insert(0, "INV", "INV")
        df_extracted.insert(1, "U", "U")
        df_extracted.insert(2, "Customer Group Code", "33001422")
        df_extracted.insert(3, "Customer Group Name", "和易 OFF")

        # Convert Minguo date to Gregorian
        def convert_minguo_date(date_str):
            try:
                if isinstance(date_str, str) and '/' in date_str:
                    year, month, day = map(int, date_str.split('/'))
                    return f"{year + 1911:04d}{month:02d}{day:02d}"
                return date_str
            except:
                return date_str

        df_extracted["Date"] = df_extracted["Date"].apply(convert_minguo_date)

        # Customer Mapping
        mapping_customer = pd.read_excel(mapping_file, sheet_name="Customer Mapping")
        mapping_customer = mapping_customer[[
            "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c")

        df_extracted = df_extracted.merge(
            mapping_customer,
            left_on="Customer Code",
            right_on="ASI_CRM_Offtake_Customer_No__c",
            how="left"
        )

        df_extracted["Customer Code"] = df_extracted["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        df_extracted.drop(columns=["ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # SKU Mapping
        mapping_sku = pd.read_excel(mapping_file, sheet_name="SKU Mapping")
        mapping_sku = mapping_sku[[
            "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"
        ]].drop_duplicates(subset="ASI_CRM_Offtake_Product__c")

        df_extracted = df_extracted.merge(
            mapping_sku,
            left_on="Product Code",
            right_on="ASI_CRM_Offtake_Product__c",
            how="left"
        )

        product_index = df_extracted.columns.get_loc("Product Code")
        df_extracted.insert(product_index, "PRT Product Code", df_extracted["ASI_CRM_SKU_Code__c"].astype(str).str.strip())
        df_extracted.drop(columns=["ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c"], inplace=True)

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_extracted)

        output_filename = "33001422_transformation.xlsx"
        df_extracted.to_excel(output_filename, index=False, header=False)

        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)

elif transformation_choice == "30010017 正興(振興)":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="zhengxing_raw")
    mapping_file = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="zhengxing_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # --- Load raw (Sheet4) ---
        df_raw = pd.read_excel(raw_data_file, sheet_name="Sheet4", header=None)

        def to_int(x):
            try:
                if pd.isna(x): return 0
                return int(float(str(x).strip()))
            except:
                return 0

        records = []
        current_product_code = None
        current_product_name = None
        doc_date = None  # YYYYMMDD (Gregorian)

        for i in range(len(df_raw)):
            row = df_raw.iloc[i]
            c0 = str(row[0]).strip() if pd.notna(row[0]) else ""
            c1 = str(row[1]).strip() if pd.notna(row[1]) else ""
            c2 = row[2] if len(row) > 2 else None   # 銷貨數量
            c3 = row[3] if len(row) > 3 else None   # 退貨數量
            c4 = row[4] if len(row) > 4 else None   # 合計數量

            # Banner date: "貨單日期: 114/07/28 至 114/07/31"  -> use right-side date
            if "貨單日期" in c0 and "至" in c0:
                m = re.search(r"至\s*([0-9]{3})/([0-9]{2})/([0-9]{2})", c0)
                if m:
                    y, mm, dd = map(int, m.groups())
                    doc_date = f"{y + 1911:04d}{mm:02d}{dd:02d}"
                continue

            # Product header: "貨品編號:G0003  貨品名稱:BEEFEATER GIN"
            if "貨品編號" in c0:
                m = re.search(r"貨品編號[:：]\s*([A-Za-z0-9\-]+)", c0)
                n = re.search(r"貨品名稱[:：]\s*(.+)", c0)
                if m: current_product_code = m.group(1).strip()
                if n: current_product_name = n.group(1).strip()
                continue

            # Skip subtotal/total lines
            if c0.startswith("小計") or c0.startswith("合計"):
                continue

            # Data rows: 客戶編號 | 客戶簡稱 | 銷貨數量 | 退貨數量 | 合計數量
            if c0 and c1 and (pd.notna(c2) or pd.notna(c3) or pd.notna(c4)):
                sales = to_int(c2)
                returns = to_int(c3)
                qty = sales - returns
                if qty == 0:
                    # Fallback to 合計數量 if both sides are blank / 0
                    qty = to_int(c4)
                if qty == 0:
                    continue

                records.append([
                    "INV", "U", "30010017", "正興(振興)",
                    c0, c1, doc_date, None,                 # PRT_Product_Code (None if unmapped)
                    current_product_code, current_product_name, qty
                ])

        if not records:
            st.warning("No transactional rows were parsed. Please verify the sheet layout.")
            st.stop()

        df_parsed = pd.DataFrame(records, columns=[
            "Type","Action","GroupCode","GroupName",
            "CustomerCode","CustomerName","Date",
            "PRT_Product_Code","ProductCode","ProductName","Quantity"
        ])

        # --- Load mappings (do not force replacements) ---
        xls_map = pd.ExcelFile(mapping_file)
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_map["ASI_CRM_Offtake_Customer_No__c"] = cust_map["ASI_CRM_Offtake_Customer_No__c"].astype(str).str.strip()
        sku_map["ASI_CRM_Offtake_Product__c"] = sku_map["ASI_CRM_Offtake_Product__c"].astype(str).str.strip()

        # Customer mapping: replace with JDE when available; otherwise leave BLANK
        df_parsed = df_parsed.merge(
            cust_map[["ASI_CRM_Offtake_Customer_No__c","ASI_CRM_JDE_Cust_No_Formula__c"]]
                .drop_duplicates(subset="ASI_CRM_Offtake_Customer_No__c"),
            left_on="CustomerCode", right_on="ASI_CRM_Offtake_Customer_No__c", how="left"
        )
        df_parsed["CustomerCode"] = (
            df_parsed["ASI_CRM_JDE_Cust_No_Formula__c"]
            .fillna("")
            .astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
        df_parsed.drop(columns=["ASI_CRM_Offtake_Customer_No__c","ASI_CRM_JDE_Cust_No_Formula__c"], inplace=True)

        # SKU mapping: fill PRT_Product_Code when available; else leave as NaN (do NOT force)
        df_parsed = df_parsed.merge(
            sku_map[["ASI_CRM_Offtake_Product__c","ASI_CRM_SKU_Code__c"]],
            left_on="ProductCode", right_on="ASI_CRM_Offtake_Product__c", how="left"
        )
        df_parsed["PRT_Product_Code"] = df_parsed["ASI_CRM_SKU_Code__c"]
        df_parsed.drop(columns=["ASI_CRM_Offtake_Product__c","ASI_CRM_SKU_Code__c"], inplace=True)

        # --- De-duplicate exact duplicates (keep first) ---
        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_final = df_parsed.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # Final order (no headers / no index on export)
        df_final = df_final[[
            "Type","Action","GroupCode","GroupName",
            "CustomerCode","CustomerName","Date",
            "PRT_Product_Code","ProductCode","ProductName","Quantity"
        ]]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        # Export: no headers, no index
        output_filename = "30010017 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )


elif transformation_choice == "30010031 廣茂隆(八條)":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="gml_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="gml_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---- Load raw (single sheet like '0728-0731') ----
        xls = pd.ExcelFile(raw_data_file)
        sheet_name = xls.sheet_names[0]
        df_raw = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None)

        # First row is header row
        df_raw.columns = df_raw.iloc[0]
        df_raw = df_raw.iloc[1:].reset_index(drop=True)

        # Standardize columns
        rename_map = {"客戶": "CustomerCode", "客戶名稱": "CustomerName",
                      "品號": "ProductCode", "品名規格": "ProductName", "銷量": "Quantity"}
        df = df_raw.rename(columns=rename_map)[["CustomerCode","CustomerName","ProductCode","ProductName","Quantity"]]

        # Quantity -> int, drop zeros
        def to_int(x):
            try:
                if pd.isna(x): return 0
                return int(float(str(x).strip()))
            except:
                return 0

        df["Quantity"] = df["Quantity"].apply(to_int)
        df = df[df["Quantity"] != 0].copy()

        # Normalize codes to strings (strip possible .0)
        for col in ["CustomerCode","ProductCode"]:
            df[col] = df[col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

        # ---- Date from sheet name: e.g. '0728-0731' -> use end date '0731' -> default year 2025 ----
        m = re.match(r"^(\d{2})(\d{2})-(\d{2})(\d{2})$", sheet_name)
        if m:
            mm_end, dd_end = m.group(3), m.group(4)
            year = "2025"  # adjust if you want a parameterized year
            date_val = f"{year}{mm_end}{dd_end}"
        else:
            # fallback: keep blank if the sheet name isn’t in the expected pattern
            date_val = None
        df["Date"] = date_val

        # ---- Load mappings (dtype=str), FILTERED to this wholesaler (30010031) ----
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_map = cust_map[
            cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30010031"
        ].copy()
        sku_map = sku_map[
            sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30010031"
        ].copy()

        cust_map["ASI_CRM_Offtake_Customer_No__c"] = cust_map["ASI_CRM_Offtake_Customer_No__c"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Offtake_Product__c"] = sku_map["ASI_CRM_Offtake_Product__c"].astype(str).str.strip()

        # ---- Customer mapping (non-forced): use JDE when present, else keep original ----
        df = df.merge(
            cust_map[["ASI_CRM_Offtake_Customer_No__c","ASI_CRM_JDE_Cust_No_Formula__c"]],
            left_on="CustomerCode", right_on="ASI_CRM_Offtake_Customer_No__c", how="left"
        )
        df["CustomerCode"] = (
            df["ASI_CRM_JDE_Cust_No_Formula__c"]
            .fillna("")  # <- key change: no fallback to external code
            .astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )

        # ---- SKU mapping (non-forced): fill PRT SKU when present, else leave NaN ----
        df = df.merge(
            sku_map[["ASI_CRM_Offtake_Product__c","ASI_CRM_SKU_Code__c"]],
            left_on="ProductCode", right_on="ASI_CRM_Offtake_Product__c", how="left"
        )
        df["PRT_Product_Code"] = df["ASI_CRM_SKU_Code__c"]
        df.drop(columns=["ASI_CRM_Offtake_Product__c","ASI_CRM_SKU_Code__c"], inplace=True)

        # ---- Add metadata columns and order ----
        df.insert(0, "Type", "INV")
        df.insert(1, "Action", "U")
        df.insert(2, "GroupCode", "30010031")
        df.insert(3, "GroupName", "廣茂隆(八條)")

        df_final = df[[
            "Type","Action","GroupCode","GroupName",
            "CustomerCode","CustomerName","Date",
            "PRT_Product_Code","ProductCode","ProductName","Quantity"
        ]].copy()

        # ---- De-duplicate exact duplicates ----
        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---- Preview + Export (NO headers, NO index) ----
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "30010031 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )
elif transformation_choice == "30020016 日嵩":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="risong_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="risong_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---------- 1) Load & detect header row ----------
        xls = pd.ExcelFile(raw_data_file)
        sheet = xls.sheet_names[0]  # expected 'AAA'
        raw = pd.read_excel(raw_data_file, sheet_name=sheet, header=None)

        header_row_idx = None
        for i in range(min(15, len(raw))):
            row_vals = raw.iloc[i].astype(str).tolist()
            if ("貨號" in row_vals[0]) and ("客戶" in (row_vals[2] if len(row_vals) > 2 else "")):
                header_row_idx = i
                break
        if header_row_idx is None:
            header_row_idx = 3  # fallback if layout shifts

        df = pd.read_excel(raw_data_file, sheet_name=sheet, header=None, skiprows=header_row_idx)
        df.columns = ["ProductCode","ProductName","CustomerCode","CustomerName","FreeQty","SalesQty","ReturnQty","NetQty"]

        # remove lingering column header row if any
        df = df[df["CustomerCode"] != "客戶"].copy()

        # ---------- 2) Numeric casting & keep Net != 0 ----------
        for col in ["FreeQty","SalesQty","ReturnQty","NetQty"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
        df = df[df["NetQty"] != 0].copy()

        # ---------- 3) Date from banner line (use END of range) ----------
        row2 = " ".join([str(x) for x in raw.iloc[2].tolist() if pd.notna(x)])
        dates = re.findall(r'(\d{4})/(\d{2})/(\d{2})', row2)
        date_val = None
        if dates:
            y, m, d = dates[-1]                       # end date
            date_val = f"{y}{m}{d}"
        df["Date"] = date_val

        # ---------- 4) Key normalization ----------
        def norm_cust(s: str) -> str:
            s = str(s).strip()
            s = re.sub(r'\.0$', '', s)
            s = s.upper().replace(' ', '')
            return s

        def norm_sku(s: str) -> str:
            s = str(s).strip().upper().replace(' ', '')
            # full-width letters -> half-width (defensive)
            s = s.replace('Ｌ','L').replace('Ａ','A').replace('Ｂ','B')
            return s

        df["CustomerCode_norm"] = df["CustomerCode"].apply(norm_cust)
        df["ProductCode_norm"]  = df["ProductCode"].apply(norm_sku)

        # ---------- 5) Load mappings ----------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        # filtered (preferred) + global fallback
        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)=="30020016"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)=="30020016"].copy()

        # prep normalized key/value frames
        def prep_cust(dfm):
            out = dfm.copy()
            out["key"] = (out["ASI_CRM_Offtake_Customer_No__c"]
                          .astype(str).str.strip().str.upper()
                          .str.replace(r"\.0$", "", regex=True).str.replace(' ', '', regex=False))
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm):
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper().str.replace(' ', '', regex=False)
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        cust_f_kv   = prep_cust(cust_f)
        cust_all_kv = prep_cust(cust_map)
        sku_f_kv    = prep_sku(sku_f)
        sku_all_kv  = prep_sku(sku_map)

        # keep only keys with a SINGLE unique target (avoid fan-out duplicates)
        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq_keys = g[g["n"]==1]["key"]
            return kv[kv["key"].isin(uniq_keys)].drop_duplicates(subset=["key"], keep="first")

        cust_f_unique   = unique_only(cust_f_kv)
        cust_all_unique = unique_only(cust_all_kv)
        sku_f_unique    = unique_only(sku_f_kv)
        sku_all_unique  = unique_only(sku_all_kv)

        # dictionaries for fast, one-to-one lookups
        cust_f_dict   = dict(zip(cust_f_unique["key"],   cust_f_unique["val"]))
        cust_all_dict = dict(zip(cust_all_unique["key"], cust_all_unique["val"]))
        sku_f_dict    = dict(zip(sku_f_unique["key"],    sku_f_unique["val"]))
        sku_all_dict  = dict(zip(sku_all_unique["key"],  sku_all_unique["val"]))

        # ---------- 6) Apply mapping WITHOUT forcing, and avoid many-to-one fan-out ----------
        jde_from_filtered = df["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df["CustomerCode_norm"].map(cust_all_dict)
        df["CustomerCode_final"] = (
            jde_from_filtered.fillna(jde_from_global).fillna(df["CustomerCode_norm"])
            .astype(str).str.replace(r"\.0$", "", regex=True)
        )

        prt_from_filtered = df["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df["ProductCode_norm"].map(sku_all_dict)
        df["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # leave NaN if still missing

        # ---------- 7) Assemble final & dedup ----------
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020016",
            "GroupName": "日嵩",
            "CustomerCode": df["CustomerCode_final"],
            "CustomerName": df["CustomerName"],
            "Date": df["Date"],
            "PRT_Product_Code": df["PRT_Product_Code"],
            "ProductCode": df["ProductCode_norm"],
            "ProductName": df["ProductName"],
            "Quantity": df["NetQty"].astype(int)
        })

        # remove exact duplicates; this plus unique-only mapping prevents the “same row repeated 5 times” issue
        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---------- 8) Preview & export (no headers / no index) ----------
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "30020016 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)
          
elif transformation_choice == "30020027 榮好(實儀)":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="ronghao_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="ronghao_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---- 1) Load primary sheet (e.g., '20250317-20250322') ----
        xls = pd.ExcelFile(raw_data_file)
        sheet = xls.sheet_names[0]
        df_raw = pd.read_excel(raw_data_file, sheet_name=sheet, header=None)

        # First row is the header row
        df_raw.columns = df_raw.iloc[0]
        df_body = df_raw.iloc[1:].reset_index(drop=True)

        # Standardize expected columns
        rename_map = {
            "客戶代碼": "CustomerCode",
            "客戶名稱": "CustomerName",
            "產品代號": "ProductCode",
            "品名規格": "ProductName",
            "銷量":   "Quantity",
        }
        df = df_body.rename(columns=rename_map)[["CustomerCode","CustomerName","ProductCode","ProductName","Quantity"]]

        # 2) Forward-fill customer code/name for continuation rows
        df["CustomerCode"] = df["CustomerCode"].ffill()
        df["CustomerName"] = df["CustomerName"].ffill()

        # 3) Quantity -> int, keep only non-zero
        df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
        df = df[df["Quantity"] != 0].copy()

        # 4) Normalize keys
        def norm_cust(s: str) -> str:
            s = str(s).strip()
            s = re.sub(r"\.0$", "", s)
            return s

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        df["CustomerCode_norm"] = df["CustomerCode"].apply(norm_cust)
        df["ProductCode_norm"]  = df["ProductCode"].apply(norm_sku)

        # 5) Date from sheet name: 'YYYYMMDD-YYYYMMDD' -> use END date
        m = re.match(r"^(\d{8})-(\d{8})$", sheet)
        date_val = m.group(2) if m else None
        df["Date"] = date_val

        # ---- 6) Load mappings (dtype=str) ----
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        # Filter to this wholesaler (preferred), but keep global as fallback
        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020027"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020027"].copy()

        # Prepare normalized key/value frames
        def prep_cust(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Customer_No__c"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq_keys = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq_keys)].drop_duplicates(subset=["key"], keep="first")

        cust_f_kv    = unique_only(prep_cust(cust_f))
        cust_all_kv  = unique_only(prep_cust(cust_map))
        sku_f_kv     = unique_only(prep_sku(sku_f))
        sku_all_kv   = unique_only(prep_sku(sku_map))

        cust_f_dict   = dict(zip(cust_f_kv["key"],   cust_f_kv["val"]))
        cust_all_dict = dict(zip(cust_all_kv["key"], cust_all_kv["val"]))
        sku_f_dict    = dict(zip(sku_f_kv["key"],    sku_f_kv["val"]))
        sku_all_dict  = dict(zip(sku_all_kv["key"],  sku_all_kv["val"]))

        # 7) Apply mapping WITHOUT fan-out (unique-only); if ambiguous or missing, keep external
        jde_from_filtered = df["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df["CustomerCode_norm"].map(cust_all_dict)
        df["CustomerCode_final"] = (
            jde_from_filtered.fillna(jde_from_global).fillna(df["CustomerCode_norm"])
            .astype(str).str.replace(r"\.0$", "", regex=True)
        )

        prt_from_filtered = df["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df["ProductCode_norm"].map(sku_all_dict)
        df["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # leave NaN if still missing

        # 8) Assemble final ordered frame
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020027",
            "GroupName": "榮好(實儀)",
            "CustomerCode": df["CustomerCode_final"],
            "CustomerName": df["CustomerName"],
            "Date": df["Date"],
            "PRT_Product_Code": df["PRT_Product_Code"],
            "ProductCode": df["ProductCode_norm"],
            "ProductName": df["ProductName"],
            "Quantity": df["Quantity"],
        })

        # 9) De-duplicate exact duplicates
        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # 10) Preview + Export (NO headers / NO index)
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "30020027 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )

elif transformation_choice == "30020180 暐倫 OFF":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="weilen_off_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="weilen_off_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---------- 1) Load raw (first row is header) ----------
        xls = pd.ExcelFile(raw_data_file)
        sheet = xls.sheet_names[0]  # e.g., '工作表1'
        df_raw = pd.read_excel(raw_data_file, sheet_name=sheet, header=None)
        df_raw.columns = df_raw.iloc[0]
        df = df_raw.iloc[1:].reset_index(drop=True)

        # Standardize expected columns (defensive rename)
        rename = {
            "銷貨單號": "DocumentNo",
            "銷貨日期": "DateRaw",
            "客戶代號": "CustomerCode",
            "客戶名稱": "CustomerName",
            "產品編號": "ProductCode",
            "產品名稱": "ProductName",
            "數量":    "Quantity",
        }
        df = df.rename(columns=rename)

        # Keep only needed columns if they exist
        cols_needed = ["DateRaw","CustomerCode","CustomerName","ProductCode","ProductName","Quantity"]
        df = df[[c for c in cols_needed if c in df.columns]].copy()

        # ---------- 2) Date → YYYYMMDD ----------
        def to_yyyymmdd(x):
            # try pandas first (handles Excel serials, datetime, strings)
            try:
                return pd.to_datetime(x).strftime("%Y%m%d")
            except Exception:
                s = str(x)
                m = re.search(r"(\d{4})[/-](\d{2})[/-](\d{2})", s)
                if m:
                    return f"{m.group(1)}{m.group(2)}{m.group(3)}"
                return None

        df["Date"] = df["DateRaw"].apply(to_yyyymmdd) if "DateRaw" in df.columns else None

        # ---------- 3) Quantity → int; keep only non-zero ----------
        df["Quantity"] = pd.to_numeric(df.get("Quantity", 0), errors="coerce").fillna(0).astype(int)
        df = df[df["Quantity"] != 0].copy()

        # ---------- 4) Normalize keys ----------
        def norm_cust(s: str) -> str:
            s = str(s).strip()
            s = re.sub(r"\.0$", "", s)
            return s

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        df["CustomerCode_norm"] = df.get("CustomerCode", "").apply(norm_cust)
        df["ProductCode_norm"]  = df.get("ProductCode", "").apply(norm_sku)

        # ---------- 5) Load mappings ----------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        # Prefer mappings filtered to this wholesaler; fallback to global
        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020180"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020180"].copy()

        def prep_cust(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = (
                out["ASI_CRM_Offtake_Customer_No__c"]
                .astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            )
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = (
                out["ASI_CRM_Offtake_Product__c"]
                .astype(str).str.strip().str.upper()
            )
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        # keep only one-to-one keys to avoid fan-out duplicates
        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq)].drop_duplicates(subset=["key"], keep="first")

        cust_f_kv    = unique_only(prep_cust(cust_f))
        cust_all_kv  = unique_only(prep_cust(cust_map))
        sku_f_kv     = unique_only(prep_sku(sku_f))
        sku_all_kv   = unique_only(prep_sku(sku_map))

        cust_f_dict   = dict(zip(cust_f_kv["key"],   cust_f_kv["val"]))
        cust_all_dict = dict(zip(cust_all_kv["key"], cust_all_kv["val"]))
        sku_f_dict    = dict(zip(sku_f_kv["key"],    sku_f_kv["val"]))
        sku_all_dict  = dict(zip(sku_all_kv["key"],  sku_all_kv["val"]))

        # ---------- 6) Apply mapping (non-forced, unique-only) ----------
        jde_from_filtered = df["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df["CustomerCode_norm"].map(cust_all_dict)
        df["CustomerCode_final"] = (
            jde_from_filtered.fillna(jde_from_global).fillna(df["CustomerCode_norm"])
            .astype(str).str.replace(r"\.0$", "", regex=True)
        )

        prt_from_filtered = df["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df["ProductCode_norm"].map(sku_all_dict)
        df["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # leave NaN if still missing

        # ---------- 7) Assemble final ordered frame ----------
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020180",
            "GroupName": "暐倫 OFF",
            "CustomerCode": df["CustomerCode_final"],
            "CustomerName": df.get("CustomerName", ""),
            "Date": df["Date"],
            "PRT_Product_Code": df["PRT_Product_Code"],
            "ProductCode": df["ProductCode_norm"],
            "ProductName": df.get("ProductName", ""),
            "Quantity": df["Quantity"],
        })

        # ---------- 8) De-duplicate exact duplicates ----------
        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---------- 9) Preview & export (no headers / no index) ----------
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "30020180 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )
elif transformation_choice == "30020203 玄星 OFF":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="xuanxing_off_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="xuanxing_off_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------------------
        # Helpers
        # ---------------------------
        def minguo_to_yyyymmdd(val):
            if pd.isna(val):
                return None
            s = str(val).strip()
            try:
                y, m, d = s.split("/")
                y, m, d = int(y), int(m), int(d)
                if y < 1911: y += 1911
                return f"{y:04d}{m:02d}{d:02d}"
            except Exception:
                try:
                    return pd.to_datetime(s).strftime("%Y%m%d")
                except Exception:
                    return None

        def to_int(x):
            try: return int(float(x))
            except: return 0

        def norm_cust(s: str) -> str:
            s = str(s).strip().upper().replace(" ", "")
            return re.sub(r"\.0$", "", s)

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        # ---------------------------
        # 1) Load all monthly sheets named like 11401..11412
        # ---------------------------
        xls = pd.ExcelFile(raw_data_file)
        month_sheets = [s for s in xls.sheet_names if re.fullmatch(r"\d{5}", s)]

        def extract_month(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None)

            # find header row where C="客戶編號", D="客戶簡稱", E="產品編號"
            header_idx = None
            for i in range(min(25, len(df))):
                c = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
                d = str(df.iat[i, 3]).strip() if df.shape[1] > 3 else ""
                e = str(df.iat[i, 4]).strip() if df.shape[1] > 4 else ""
                if c == "客戶編號" and d == "客戶簡稱" and e == "產品編號":
                    header_idx = i
                    break
            if header_idx is None:
                return pd.DataFrame()

            rows = []
            for r in range(header_idx + 1, len(df)):
                if str(df.iat[r, 0]).strip() == "合計":
                    break

                date_cell = df.iat[r, 0]
                cust_code = df.iat[r, 2]
                cust_name = df.iat[r, 3]
                prod_code = df.iat[r, 4]
                prod_name = df.iat[r, 5]
                sales_qty = df.iat[r, 6]
                free_qty  = df.iat[r, 7] if df.shape[1] > 7 else 0

                if pd.isna(prod_code) and pd.isna(prod_name) and pd.isna(cust_code):
                    continue

                qty = to_int(sales_qty) + to_int(free_qty)
                if qty == 0:
                    continue

                rows.append({
                    "Date": minguo_to_yyyymmdd(date_cell),
                    "CustomerCode": cust_code,
                    "CustomerName": cust_name,
                    "ProductCode": prod_code,
                    "ProductName": prod_name,
                    "Quantity": qty
                })
            return pd.DataFrame(rows)

        df_all = pd.concat([extract_month(s) for s in month_sheets], ignore_index=True)

        if df_all.empty:
            st.warning("No valid rows found across monthly tabs.")
            st.stop()

        # ---------------------------
        # 2) Normalize + month key
        # ---------------------------
        df_all["CustomerCode_norm"] = df_all["CustomerCode"].apply(norm_cust)
        df_all["ProductCode_norm"]  = df_all["ProductCode"].apply(norm_sku)
        df_all["Month"] = df_all["Date"].astype(str).str[:6]  # YYYYMM

        # ---------------------------
        # 3) Load mappings (unique-only; prefer filtered, then global)
        # ---------------------------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020203"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020203"].copy()

        def prep_cust(dfm):
            out = dfm.copy()
            out["key"] = (out["ASI_CRM_Offtake_Customer_No__c"].astype(str)
                          .str.strip().str.upper().str.replace(r"\.0$", "", regex=True).str.replace(" ", "", regex=False))
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm):
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq)].drop_duplicates(subset=["key"], keep="first")

        cust_f_dict   = dict(zip(unique_only(prep_cust(cust_f))["key"],   unique_only(prep_cust(cust_f))["val"]))
        cust_all_dict = dict(zip(unique_only(prep_cust(cust_map))["key"], unique_only(prep_cust(cust_map))["val"]))
        sku_f_dict    = dict(zip(unique_only(prep_sku(sku_f))["key"],     unique_only(prep_sku(sku_f))["val"]))
        sku_all_dict  = dict(zip(unique_only(prep_sku(sku_map))["key"],   unique_only(prep_sku(sku_map))["val"]))

        # ---------------------------
        # 4) Apply mapping
        #     CHANGE: if no mapping, leave CustomerCode blank (not original)
        # ---------------------------
        jde_from_filtered = df_all["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df_all["CustomerCode_norm"].map(cust_all_dict)
        mapped_jde        = jde_from_filtered.combine_first(jde_from_global)
        df_all["CustomerCode_final"] = mapped_jde.fillna("")  # blank when unmapped

        prt_from_filtered = df_all["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df_all["ProductCode_norm"].map(sku_all_dict)
        df_all["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # leave NaN if missing

        # ---------------------------
        # 5) Build final frame (all months), de-dupe
        # ---------------------------
        df_all_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020203",
            "GroupName": "玄星 OFF",
            "CustomerCode": df_all["CustomerCode_final"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
            "Month": df_all["Month"],
        })

        dedup_keys = ["GroupCode","CustomerCode","Date","ProductCode","Quantity"]
        df_all_final = df_all_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---------------------------
        # 6) UI: Toggle by Month (📅)
        # ---------------------------
        available_months = sorted(df_all_final["Month"].dropna().astype(str).unique().tolist())
        month_filter = st.radio("📅 Filter by Month:", ["All"] + available_months, index=0)

        if month_filter != "All":
            df_view = df_all_final[df_all_final["Month"] == month_filter].copy()
        else:
            df_view = df_all_final.copy()

        # Drop helper Month column from display/export
        df_view = df_view[[
            "Type","Action","GroupCode","GroupName",
            "CustomerCode","CustomerName","Date",
            "PRT_Product_Code","ProductCode","ProductName","Quantity"
        ]]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_view)

        # ---------------------------
        # 7) Export selection (no headers / no index)
        # ---------------------------
        out_name = "30020203_玄星OFF_all_months.xlsx" if month_filter == "All" else f"30020203_玄星OFF_{month_filter}.xlsx"
        df_view.to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button(label="📥 Download Selected Month", data=f, file_name=out_name)

elif transformation_choice == "30020216 久悅貿易":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="jiuyue_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="jiuyue_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---------- 1) Pick a YYYYMM sheet if present ----------
        xls = pd.ExcelFile(raw_data_file)
        month_like = [s for s in xls.sheet_names if re.fullmatch(r"\d{6}", s)]
        sheet = month_like[0] if month_like else xls.sheet_names[0]
        df_raw = pd.read_excel(raw_data_file, sheet_name=sheet, header=None)

        # ---------- 2) Helpers ----------
        def minguo_to_yyyymmdd(s):
            if pd.isna(s): 
                return None
            try:
                y, m, d = str(s).strip().split("/")
                y, m, d = int(y), int(m), int(d)
                if y < 1911: y += 1911
                return f"{y:04d}{m:02d}{d:02d}"
            except Exception:
                return None

        def parse_customer(line: str):
            s = re.sub(r'[\u200b\ufeff]', '', line)
            s = re.sub(r'\s+', ' ', s)
            m = re.search(r'客戶簡稱[:：]\s*([A-Z0-9\-]+)\s+(.+?)(?:\s+電\s*話|$)', s)
            if m:
                return m.group(1).strip(), m.group(2).strip()
            return None, None

        def last_code_token(name: str):
            toks = str(name).strip().split()
            if toks:
                tail = toks[-1]
                if re.fullmatch(r"[A-Za-z0-9\-]+", tail):
                    return tail
            return ""

        # ---------- 3) Extract rows ----------
        records = []
        current_cust_code = None
        current_cust_name = None
        current_date = None

        for i in range(len(df_raw)):
            c0 = df_raw.iat[i, 0]
            c1 = df_raw.iat[i, 1] if df_raw.shape[1] > 1 else None   # ← Document No
            c2 = df_raw.iat[i, 2] if df_raw.shape[1] > 2 else None   # Name
            c3 = df_raw.iat[i, 3] if df_raw.shape[1] > 3 else None   # Qty

            # New customer block
            if isinstance(c0, str) and c0.strip().startswith("客戶簡稱"):
                code, name = parse_customer(c0)
                if code:
                    current_cust_code = code
                    current_cust_name = name
                continue

            # Skip headers/totals
            if str(c0).strip() in ["單據日期", "合    計：", "合計", "合    計:"]:
                continue

            # Date line
            if isinstance(c0, str) and re.match(r"^\d{3}/\d{2}/\d{2}$", c0.strip()):
                current_date = minguo_to_yyyymmdd(c0.strip())

            # Product rows: robust qty parse (handles "3.00" strings)
            prod_name = c2 if isinstance(c2, str) else (str(c2) if pd.notna(c2) else None)
            qty = pd.to_numeric(c3, errors="coerce")
            if prod_name and pd.notna(qty):
                qty = int(float(qty))
                if qty == 0:
                    continue
                prod_code = last_code_token(prod_name)
                records.append({
                    "DocumentNo": c1,                         # NEW
                    "CustomerCode_ext": current_cust_code,    # keep external for de-dup only
                    "CustomerName": current_cust_name,
                    "Date": current_date,
                    "ProductCode": prod_code,
                    "ProductName": prod_name,
                    "Quantity": qty
                })

        if not records:
            st.warning("No valid rows found.")
            st.stop()

        df_rec = pd.DataFrame(records)

        # ---------- 4) Normalize keys ----------
        def norm_cust(s: str) -> str:
            s = str(s).strip().upper().replace(" ", "")
            return re.sub(r"\.0$", "", s)

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        df_rec["CustomerCode_norm"] = df_rec["CustomerCode_ext"].apply(norm_cust)
        df_rec["ProductCode_norm"]  = df_rec["ProductCode"].apply(norm_sku)

        # ---------- 5) Load mappings (unique-only; prefer filtered, then global) ----------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020216"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30020216"].copy()

        def prep_cust(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = (out["ASI_CRM_Offtake_Customer_No__c"].astype(str)
                          .str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
                          .str.replace(" ", "", regex=False))
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm: pd.DataFrame) -> pd.DataFrame:
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq)].drop_duplicates(subset=["key"], keep="first")

        cust_f_dict   = dict(zip(unique_only(prep_cust(cust_f))["key"],   unique_only(prep_cust(cust_f))["val"]))
        cust_all_dict = dict(zip(unique_only(prep_cust(cust_map))["key"], unique_only(prep_cust(cust_map))["val"]))
        sku_f_dict    = dict(zip(unique_only(prep_sku(sku_f))["key"],     unique_only(prep_sku(sku_f))["val"]))
        sku_all_dict  = dict(zip(unique_only(prep_sku(sku_map))["key"],   unique_only(prep_sku(sku_map))["val"]))

        # ---------- 6) Apply mapping
        # Leave CustomerCode BLANK when unmapped (per your requirement)
        # ----------
        jde_from_filtered = df_rec["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df_rec["CustomerCode_norm"].map(cust_all_dict)
        df_rec["CustomerCode_final"] = jde_from_filtered.combine_first(jde_from_global).fillna("")

        prt_from_filtered = df_rec["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df_rec["ProductCode_norm"].map(sku_all_dict)
        df_rec["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # may stay NaN

        # ---------- 7) Build output ----------
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020216",
            "GroupName": "久悅貿易",
            "CustomerCode": df_rec["CustomerCode_final"],          # blank if unmapped
            "CustomerName": df_rec["CustomerName"],
            "Date": df_rec["Date"],
            "PRT_Product_Code": df_rec["PRT_Product_Code"],
            "ProductCode": df_rec["ProductCode_norm"],
            "ProductName": df_rec["ProductName"],
            "Quantity": df_rec["Quantity"].astype(int),
            "DocumentNo": df_rec["DocumentNo"],                    # helper (not exported)
            "_custKey": df_rec["CustomerCode_final"].mask(
                df_rec["CustomerCode_final"].eq(""), df_rec["CustomerCode_norm"]
            ),  # use external code only for de-dup key
        })

        # ---------- 8) De-dup: keep distinct lines (uses doc no + product name + custKey) ----------
        dedup_keys = ["GroupCode","_custKey","Date","ProductCode","ProductName","Quantity","DocumentNo"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # drop helper columns from export
        df_export = df_final.drop(columns=["DocumentNo","_custKey"])

        # ---------- 9) Preview + Export (no headers / no index) ----------
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_export)

        output_filename = "30020216 transformation.xlsx"
        df_export.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(label="📥 Download Processed File", data=f, file_name=output_filename)
elif transformation_choice == "30030061 合歡 OFF":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="hehuan_off_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="hehuan_off_mapping")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------------------
        # Helpers
        # ---------------------------
        def minguo_to_yyyymmdd(s):
            if pd.isna(s): 
                return None
            s = str(s).strip()
            try:
                y, m, d = s.split("/")
                y, m, d = int(y), int(m), int(d)
                if y < 1911: y += 1911
                return f"{y:04d}{m:02d}{d:02d}"
            except Exception:
                # graceful fallback (already Gregorian / Excel date)
                try:
                    return pd.to_datetime(s).strftime("%Y%m%d")
                except Exception:
                    return None

        def clean_name(s):
            if pd.isna(s): return ""
            s = str(s).strip()
            s = re.sub(r'^[\[\【]\s*', '', s)  # leading bracket
            s = re.sub(r'\s*[\]\】]$', '', s)  # trailing bracket
            return s

        def norm_cust(s: str) -> str:
            s = str(s).strip().upper().replace(" ", "")
            return re.sub(r"\.0$", "", s)

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        # ---------------------------
        # 1) Load first sheet
        # ---------------------------
        xls = pd.ExcelFile(raw_data_file)
        sheet = xls.sheet_names[0]
        df = pd.read_excel(raw_data_file, sheet_name=sheet, header=None)

        # ---------------------------
        # 2) Parse: walk "產品編號" blocks, sum qty per document/customer/product
        # ---------------------------
        records = []
        current_prod_code = None
        current_prod_name = None
        current_date = None
        current_doc  = None

        for i in range(len(df)):
            c0 = df.iat[i, 0]
            c1 = df.iat[i, 1] if df.shape[1] > 1 else None  # 單據號碼
            c2 = df.iat[i, 2] if df.shape[1] > 2 else None  # 客戶編號
            c3 = df.iat[i, 3] if df.shape[1] > 3 else None  # 客戶簡稱
            c5 = df.iat[i, 5] if df.shape[1] > 5 else None  # 數量

            # header row announcing a product block
            if isinstance(c0, str) and c0.strip().startswith("產品編號"):
                # product code usually in col1, product name in col3 (bracketed)
                current_prod_code = norm_sku(c1) if pd.notna(c1) else ""
                current_prod_name = clean_name(df.iat[i, 3] if df.shape[1] > 3 else "")
                continue

            # skip totals/footer
            if isinstance(c2, str) and "台幣合計" in c2:
                continue

            # date (Minguo) appears in col0; forward-fill to free lines
            if isinstance(c0, str) and re.match(r"^\d{3}/\d{2}/\d{2}$", c0.strip()):
                current_date = minguo_to_yyyymmdd(c0.strip())

            # document number forward-fill
            if isinstance(c1, str) and c1.strip():
                current_doc = c1.strip()

            # quantity is numeric in col5 (accept "3", "3.00")
            qty = pd.to_numeric(c5, errors="coerce")

            if pd.notna(qty) and current_prod_code and current_date:
                records.append({
                    "Date": current_date,
                    "DocumentNo": current_doc,
                    "CustomerCode_ext": str(c2).strip() if pd.notna(c2) else "",
                    "CustomerName": str(c3).strip() if pd.notna(c3) else "",
                    "ProductCode": current_prod_code,
                    "ProductName": current_prod_name,
                    "Quantity": int(float(qty))
                })

        if not records:
            st.warning("No valid rows found.")
            st.stop()

        df_txn = pd.DataFrame(records)

        # Combine sales + free within the same doc/customer/product/date
        group_keys = ["Date","DocumentNo","CustomerCode_ext","CustomerName","ProductCode","ProductName"]
        df_txn = df_txn.groupby(group_keys, as_index=False)["Quantity"].sum()

        # ---------------------------
        # 3) Mappings (unique-only; prefer filtered to 30030061, then global)
        # ---------------------------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30030061"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30030061"].copy()

        def prep_cust(dfm):
            out = dfm.copy()
            out["key"] = (out["ASI_CRM_Offtake_Customer_No__c"].astype(str)
                          .str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
                          .str.replace(" ", "", regex=False))
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm):
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq)].drop_duplicates(subset=["key"], keep="first")

        cust_f_dict   = dict(zip(unique_only(prep_cust(cust_f))["key"],   unique_only(prep_cust(cust_f))["val"]))
        cust_all_dict = dict(zip(unique_only(prep_cust(cust_map))["key"], unique_only(prep_cust(cust_map))["val"]))
        sku_f_dict    = dict(zip(unique_only(prep_sku(sku_f))["key"],     unique_only(prep_sku(sku_f))["val"]))
        sku_all_dict  = dict(zip(unique_only(prep_sku(sku_map))["key"],   unique_only(prep_sku(sku_map))["val"]))

        # Normalize keys then map
        df_txn["CustomerCode_norm"] = df_txn["CustomerCode_ext"].apply(norm_cust)
        df_txn["ProductCode_norm"]  = df_txn["ProductCode"].apply(norm_sku)

        jde_from_filtered = df_txn["CustomerCode_norm"].map(cust_f_dict)
        jde_from_global   = df_txn["CustomerCode_norm"].map(cust_all_dict)
        # Per your rule: leave blank if unmapped (do NOT keep external code)
        df_txn["CustomerCode_final"] = jde_from_filtered.combine_first(jde_from_global).fillna("")

        prt_from_filtered = df_txn["ProductCode_norm"].map(sku_f_dict)
        prt_from_global   = df_txn["ProductCode_norm"].map(sku_all_dict)
        df_txn["PRT_Product_Code"] = prt_from_filtered.fillna(prt_from_global)  # may remain NaN

        # ---------------------------
        # 4) Assemble output + safe de-dup (guard by DocumentNo)
        # ---------------------------
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030061",
            "GroupName": "合歡 OFF",
            "CustomerCode": df_txn["CustomerCode_final"],     # blank if unmapped
            "CustomerName": df_txn["CustomerName"],
            "Date": df_txn["Date"],
            "PRT_Product_Code": df_txn["PRT_Product_Code"],
            "ProductCode": df_txn["ProductCode_norm"],
            "ProductName": df_txn["ProductName"],
            "Quantity": df_txn["Quantity"].astype(int),
            "DocumentNo": df_txn["DocumentNo"]
        })

        dedup_keys = ["DocumentNo","CustomerCode","Date","ProductCode","ProductName","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---------------------------
        # 5) Preview & Export (no headers / no index)
        # ---------------------------
        st.write("✅ Processed Data Preview:")
        st.dataframe(df_final)

        output_filename = "30030061 transformation.xlsx"
        df_final.to_excel(output_filename, index=False, header=False)
        with open(output_filename, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=output_filename
            )

elif transformation_choice == "30030076 裕陞（分月）":
    import re
    import pandas as pd
    import streamlit as st

    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xlsx)", type=["xlsx"], key="yusheng_raw_v4")
    mapping_file  = st.file_uploader("Upload Mapping File (.xlsx)", type=["xlsx"], key="yusheng_map_v4")

    if raw_data_file is not None and mapping_file is not None:
        # ---------- Helpers ----------
        def norm_cust(s: str) -> str:
            s = str(s).strip().upper().replace(" ", "")
            return re.sub(r"\.0$", "", s)

        def norm_sku(s: str) -> str:
            return str(s).strip().upper()

        # ---------- 1) Parse ALL sheets (multi product blocks per sheet) ----------
        xls = pd.ExcelFile(raw_data_file)
        sheets = xls.sheet_names

        def extract_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None)
            if df.empty:
                return pd.DataFrame()

            rows = []
            current_prod_code, current_prod_name = "", ""
            in_table = False  # inside 日期/銷貨單號/客戶編號/客戶簡稱 grid

            def sval(r, c):
                return str(df.iat[r, c]).strip() if (df.shape[1] > c and pd.notna(df.iat[r, c])) else ""

            for r in range(len(df)):
                s0, s1, s2, s3 = sval(r, 0), sval(r, 1), sval(r, 2), sval(r, 3)

                # ---- Product header (two layouts)
                m_inline = re.match(r"^\s*(\d{6,})\s+(.+)$", s0)  # "123456 品名" in col A
                if (re.fullmatch(r"\d{6,}", s0) and s1 and ":" not in s0 and "/" not in s0) or m_inline:
                    if m_inline:
                        current_prod_code, current_prod_name = m_inline.group(1).strip(), m_inline.group(2).strip()
                    else:
                        current_prod_code, current_prod_name = s0, s1
                    in_table = False
                    continue

                # ---- Grid header
                if s0 == "日期" and s1 == "銷貨單號" and s2 == "客戶編號" and s3 == "客戶簡稱":
                    in_table = True
                    continue

                if not in_table or not current_prod_code:
                    continue

                # ---- Footer/summary lines to skip
                if any(k in s0 for k in ["合計", "小計"]):
                    continue

                # ---- Detail line
                # A: 日期  B: 銷貨單號  C: 客戶編號  D: 客戶簡稱  E: 數量
                date_cell = df.iat[r, 0] if df.shape[1] > 0 else None
                qty_cell  = df.iat[r, 4] if df.shape[1] > 4 else None

                # Some workbooks have "列印日期" at the top; DO NOT break on it—just ignore non-date cells
                try:
                    date_fmt = pd.to_datetime(date_cell).strftime("%Y%m%d")
                except Exception:
                    date_fmt = None

                qty = pd.to_numeric(qty_cell, errors="coerce")

                if date_fmt and pd.notna(qty) and float(qty) != 0:
                    rows.append({
                        "Date": date_fmt,
                        "DocumentNo": sval(r, 1),
                        "CustomerCode_ext": sval(r, 2),
                        "CustomerName": sval(r, 3),
                        "ProductCode": norm_sku(current_prod_code),
                        "ProductName": current_prod_name,
                        "Quantity": int(float(qty)),
                    })

            return pd.DataFrame(rows)

        parsed = [extract_sheet(s) for s in sheets]
        df_all = pd.concat([d for d in parsed if not d.empty], ignore_index=True)
        if df_all.empty:
            st.warning("No valid rows found across sheets.")
            st.stop()

        # Combine duplicates within same doc/customer/product/date (e.g., sales + free lines)
        group_keys = ["Date","DocumentNo","CustomerCode_ext","CustomerName","ProductCode","ProductName"]
        df_all = df_all.groupby(group_keys, as_index=False)["Quantity"].sum()

        # ---------- 2) Mappings (unique-only; prefer filtered to 30030076, then global) ----------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30030076"].copy()
        sku_f  = sku_map[ sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True) == "30030076"].copy()

        def prep_cust(dfm):
            out = dfm.copy()
            out["key"] = (out["ASI_CRM_Offtake_Customer_No__c"].astype(str)
                          .str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
                          .str.replace(" ", "", regex=False))
            out["val"] = out["ASI_CRM_JDE_Cust_No_Formula__c"].astype(str).str.strip()
            return out[["key","val"]]

        def prep_sku(dfm):
            out = dfm.copy()
            out["key"] = out["ASI_CRM_Offtake_Product__c"].astype(str).str.strip().str.upper()
            out["val"] = out["ASI_CRM_SKU_Code__c"].astype(str).str.strip()
            return out[["key","val"]]

        def unique_only(kv: pd.DataFrame) -> pd.DataFrame:
            g = kv.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = g[g["n"] == 1]["key"]
            return kv[kv["key"].isin(uniq)].drop_duplicates(subset=["key"], keep="first")

        cust_f_dict   = dict(zip(unique_only(prep_cust(cust_f))["key"],   unique_only(prep_cust(cust_f))["val"]))
        cust_all_dict = dict(zip(unique_only(prep_cust(cust_map))["key"], unique_only(prep_cust(cust_map))["val"]))
        sku_f_dict    = dict(zip(unique_only(prep_sku(sku_f))["key"],     unique_only(prep_sku(sku_f))["val"]))
        sku_all_dict  = dict(zip(unique_only(prep_sku(sku_map))["key"],   unique_only(prep_sku(sku_map))["val"]))

        df_all["CustomerCode_norm"] = df_all["CustomerCode_ext"].apply(norm_cust)
        df_all["ProductCode_norm"]  = df_all["ProductCode"].apply(norm_sku)

        jde_filtered = df_all["CustomerCode_norm"].map(cust_f_dict)
        jde_global   = df_all["CustomerCode_norm"].map(cust_all_dict)
        # Per rule: leave blank if unmapped (do NOT keep external)
        df_all["CustomerCode_final"] = jde_filtered.combine_first(jde_global).fillna("")

        prt_filtered = df_all["ProductCode_norm"].map(sku_f_dict)
        prt_global   = df_all["ProductCode_norm"].map(sku_all_dict)
        df_all["PRT_Product_Code"]   = prt_filtered.fillna(prt_global)

        # ---------- 3) Assemble output + Month key ----------
        df_all["Month"] = df_all["Date"].astype(str).str[:6]
        df_final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030076",
            "GroupName": "裕陞",
            "CustomerCode": df_all["CustomerCode_final"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
            "DocumentNo": df_all["DocumentNo"],
            "Month": df_all["Month"],
        })

        # De-dup (conservative: keep DocumentNo)
        dedup_keys = ["DocumentNo","CustomerCode","Date","ProductCode","ProductName","Quantity"]
        df_final = df_final.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)

        # ---------- 4) Multi-month selector + export ----------
        months = sorted(df_final["Month"].dropna().astype(str).unique().tolist())
        selected_months = st.multiselect("📅 Select month(s) to view/export:", options=months, default=months)

        df_view = df_final[df_final["Month"].isin(selected_months)].copy() if selected_months else df_final.head(0).copy()

        # Drop helper Month; keep DocumentNo for safety
        df_view = df_view[[
            "Type","Action","GroupCode","GroupName",
            "CustomerCode","CustomerName","Date",
            "PRT_Product_Code","ProductCode","ProductName","Quantity","DocumentNo"
        ]]

        st.write("✅ Processed Data Preview:")
        st.dataframe(df_view)

        # Filename
        if not selected_months or len(selected_months) == len(months):
            tag = "all_months"
        elif len(selected_months) <= 4:
            tag = "_".join(selected_months)
        else:
            tag = f"{selected_months[0]}_to_{selected_months[-1]}_{len(selected_months)}mo"

        out_name = f"30030076_裕陞_{tag}.xlsx"
        df_view.to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Selected Month(s)", f, file_name=out_name)

elif transformation_choice == "30010008 利多吉":
    import re
    import pandas as pd
    import streamlit as st

    # ---- Uploaders: allow both .xls and .xlsx
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls","xlsx"], key="liduoji_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls","xlsx"], key="liduoji_map")

    if raw_data_file is not None and mapping_file is not None:
        # =============== engines (.xls needs xlrd) ===============
        def pick_engine(uploaded):
            return "xlrd" if uploaded and uploaded.name.lower().endswith(".xls") else None
        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)

        # =============== helpers ===============
        def minguo_to_ymd(s):
            """114/07/31 -> 20250731; return None if not a date."""
            if s is None or (isinstance(s, float) and pd.isna(s)):
                return None
            s = str(s).strip()
            m = re.match(r"^(\d{3})/(\d{2})/(\d{2})$", s)
            if m:
                y = int(m.group(1)) + 1911
                mth = int(m.group(2))
                d = int(m.group(3))
                return f"{y:04d}{mth:02d}{d:02d}"
            # fallback: try normal datetime
            try:
                return pd.to_datetime(s).strftime("%Y%m%d")
            except Exception:
                return None

        def norm_code(s: str) -> str:
            return str(s).strip().upper().replace(" ", "").replace(".0", "")

        def unique_only_map(df, key_col, val_col, norm=lambda x: x):
            """Build key->val map keeping only keys with a single unique value."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(norm)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        # ---- header detection helpers (robust to slight shifts)
        def looks_like_header(cells):
            row = [str(c).strip() for c in cells]
            joined = "|".join(row)
            has_date = any(tok in joined for tok in ["銷貨日期", "日期"])
            has_doc  = any(tok in joined for tok in ["銷貨單號", "單據號碼"])
            has_cust = any(tok in joined for tok in ["客戶", "客戶簡稱", "客戶編號", "客戶代號"])
            has_qty  = any(tok in joined for tok in ["數量", "數量(瓶)"])
            return has_date and has_doc and has_cust and has_qty

        def find_indices(header_cells):
            """Return (date_idx, doc_idx, cust_code_idx, cust_name_idx, qty_idx) best-effort."""
            row = [str(c).strip() for c in header_cells]
            def idx_of(cands):
                for t in cands:
                    if t in row:
                        return row.index(t)
                return None
            date_idx = idx_of(["銷貨日期","日期"])
            doc_idx  = idx_of(["銷貨單號","單據號碼"])
            cust_code_idx = idx_of(["客戶編號","客戶代號"])
            cust_name_idx = idx_of(["客戶簡稱","客戶"])
            qty_idx  = idx_of(["數量","數量(瓶)"])
            # fallbacks (common layout: A,B,D,E,F)
            if date_idx is None: date_idx = 0
            if doc_idx  is None: doc_idx  = 1
            if cust_code_idx is None: cust_code_idx = 3
            if cust_name_idx is None: cust_name_idx = 4
            if qty_idx  is None: qty_idx  = 5
            return date_idx, doc_idx, cust_code_idx, cust_name_idx, qty_idx

        # =============== 1) Parse all sheets (blocks: 起訖品號 …) ===============
        xls = pd.ExcelFile(raw_data_file, engine=raw_eng)
        sheets = xls.sheet_names

        def extract_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            ncols = df.shape[1]
            recs = []
            current_code = ""
            current_name = ""
            in_grid = False
            header_idx_tuple = None  # indices for columns within the grid

            def sval(r, c):
                return str(df.iat[r, c]).strip() if (c < ncols and pd.notna(df.iat[r, c])) else ""

            def seek_name_forward(start_row: int) -> str:
                for rr in range(start_row + 1, min(start_row + 4, len(df))):
                    s3 = sval(rr, 3)
                    if s3:
                        return s3
                return ""

            for r in range(len(df)):
                # collect a window of cells for header sniffing when needed
                row_cells = [df.iat[r, c] if (c < ncols and pd.notna(df.iat[r, c])) else "" for c in range(min(12, ncols))]
                s0 = sval(r, 0)

                # ---- product header: "起訖品號：<code>" (name usually in col D)
                if isinstance(s0, str) and s0.startswith("起訖品號："):
                    current_code = s0.replace("起訖品號：", "").strip().upper()
                    # prefer same-row col D; else look forwards
                    maybe_name = sval(r, 3)
                    current_name = maybe_name if maybe_name else seek_name_forward(r)
                    in_grid = False
                    header_idx_tuple = None
                    continue

                # ---- detail grid header (robust detection)
                if looks_like_header(row_cells):
                    in_grid = True
                    header_idx_tuple = find_indices([str(x).strip() for x in row_cells])
                    continue

                if not in_grid or not current_code or header_idx_tuple is None:
                    continue

                # ---- subtotal/other non-data lines: skip (do NOT break the sheet scan)
                joined = "|".join([str(c).strip() for c in row_cells])
                if any(tag in joined for tag in ("合計", "小計")):
                    continue

                # ---- detail row using detected indices
                date_idx, doc_idx, cust_code_idx, cust_name_idx, qty_idx = header_idx_tuple
                date_cell = df.iat[r, date_idx] if date_idx < ncols else None
                doc_cell  = sval(r, doc_idx) if doc_idx < ncols else ""
                cc_cell   = sval(r, cust_code_idx) if cust_code_idx < ncols else ""
                cname_cell= sval(r, cust_name_idx) if cust_name_idx < ncols else ""

                date_ymd  = minguo_to_ymd(date_cell)
                if not date_ymd:
                    # not a detail row (blank, text, footer, next product header, etc.)
                    continue

                # quantity: first try qty_idx; else scan a few cols to the right of name
                qty_val = None
                if qty_idx < ncols:
                    qty_val = pd.to_numeric(df.iat[r, qty_idx], errors="coerce")
                if (qty_val is None) or pd.isna(qty_val):
                    for c in range(min(ncols, cust_name_idx + 1), min(ncols, cust_name_idx + 4)):
                        qv = pd.to_numeric(df.iat[r, c], errors="coerce")
                        if pd.notna(qv):
                            qty_val = qv
                            break

                if qty_val is None or float(qty_val) == 0:
                    continue

                recs.append({
                    "Sheet": sheet_name,
                    "Row": r,
                    "Date": date_ymd,
                    "DocumentNo": doc_cell,
                    "CustomerCode_ext": cc_cell,
                    "CustomerName": cname_cell,
                    "ProductCode": current_code,
                    "ProductName": current_name,
                    "Quantity": int(float(qty_val)),
                })

            return pd.DataFrame(recs)

        # --- Parse all sheets with logging, guard against empty concat
        frames = []
        parse_log = []
        for s in sheets:
            try:
                d = extract_sheet(s)
                n = 0 if d is None else len(d)
                if n:
                    frames.append(d)
                parse_log.append(f"{s}: {n} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not frames:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        df_all = pd.concat(frames, ignore_index=True)

        # Combine duplicates within the same doc/customer/product/date
        group_keys = ["Date", "DocumentNo", "CustomerCode_ext", "CustomerName", "ProductCode", "ProductName"]
        df_all = df_all.groupby(group_keys, as_index=False)["Quantity"].sum()

        # =============== 2) Mappings (unique-only; prefer filtered 30010008, then global) ===============
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str, engine=map_eng)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping", dtype=str, engine=map_eng)

        # Customer mapping
        cust_map["ASI_CRM_Mapping_Cust_No__c"] = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30010008"].copy()
        m_cust_f = unique_only_map(cust_f, "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map, "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)

        # SKU mapping
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_f = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30010008"].copy()
        m_sku_f = unique_only_map(sku_f, "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c", lambda s: str(s).strip().upper())
        m_sku_g = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c", "ASI_CRM_SKU_Code__c", lambda s: str(s).strip().upper())

        df_all["CustomerCode_norm"] = df_all["CustomerCode_ext"].map(norm_code)
        df_all["CustomerCode"] = df_all["CustomerCode_norm"].map(m_cust_f).fillna(df_all["CustomerCode_norm"].map(m_cust_g))
        # Non-forced: if unmapped, leave blank (don't keep external)
        df_all["CustomerCode"] = df_all["CustomerCode"].fillna("")

        df_all["ProductCode_norm"] = df_all["ProductCode"].str.strip().str.upper()
        df_all["PRT_Product_Code"] = df_all["ProductCode_norm"].map(m_sku_f).fillna(df_all["ProductCode_norm"].map(m_sku_g))

        # =============== 3) Assemble final output (preserve order) ===============
        df_all = df_all.sort_values(["Date", "DocumentNo"]).reset_index(drop=True)

        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30010008",
            "GroupName": "利多吉",
            "CustomerCode": df_all["CustomerCode"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
            "DocumentNo": df_all["DocumentNo"],
        })

        # ---- UI
        st.write("✅ Processed Data Preview:")
        st.dataframe(final.head(30))

        with st.expander("🔎 Parse summary (per sheet)"):
            st.code("\n".join(parse_log))

        # Export (no headers, no index)
        out_name = "30010008_利多吉_transformation.xlsx"
        final.to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30010154 亨玖":
    import re
    import calendar
    import pandas as pd
    import streamlit as st

    # ---- Uploaders: allow both .xls and .xlsx
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls","xlsx"], key="hengjiu_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls","xlsx"], key="hengjiu_map")

    if raw_data_file is not None and mapping_file is not None:
        # -------- Engines (.xls needs xlrd) --------
        def pick_engine(uploaded):
            return "xlrd" if uploaded and uploaded.name.lower().endswith(".xls") else None
        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)

        # -------- Helpers --------
        def find_period_end_ymd(frame: pd.DataFrame) -> str | None:
            """Look for '114.7' style period and return end-of-month YYYYMMDD."""
            max_r = min(len(frame), 20)
            max_c = min(frame.shape[1], 8)
            for r in range(max_r):
                for c in range(max_c):
                    val = frame.iat[r, c]
                    if pd.isna(val):
                        continue
                    s = str(val).strip()
                    m = re.match(r'^(\d{3})\.(\d{1,2})$', s)
                    if m:
                        y = int(m.group(1)) + 1911
                        mth = int(m.group(2))
                        last_day = calendar.monthrange(y, mth)[1]
                        return f"{y:04d}{mth:02d}{last_day:02d}"
            return None

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # -------- 1) Parse ALL sheets --------
        xls = pd.ExcelFile(raw_data_file, engine=raw_eng)
        sheets = xls.sheet_names

        def extract_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            # find period end date (YYYYMMDD)
            period_date = find_period_end_ymd(df)

            # locate header row (產品編號, 發票品名, 客戶/廠商簡稱, 客戶/廠商編號, 數量)
            header_row = None
            for r in range(len(df)):
                row = [str(df.iat[r, c]).strip() if (c < df.shape[1] and pd.notna(df.iat[r, c])) else "" for c in range(df.shape[1])]
                if ("產品編號" in row and "發票品名" in row and "客戶/廠商編號" in row and "數量" in row):
                    header_row = r
                    break
            if header_row is None:
                return pd.DataFrame()

            recs = []
            current_prod_code = None
            current_prod_name = None

            for r in range(header_row + 1, len(df)):
                prod_code = df.iat[r, 0] if df.shape[1] > 0 else None
                prod_name = df.iat[r, 1] if df.shape[1] > 1 else None
                cust_name = df.iat[r, 2] if df.shape[1] > 2 else None
                cust_code = df.iat[r, 3] if df.shape[1] > 3 else None
                qty_cell  = df.iat[r, 4] if df.shape[1] > 4 else None

                # New product header row
                if isinstance(prod_code, str) and prod_code.strip():
                    current_prod_code = norm_sku(prod_code)
                    current_prod_name = str(prod_name).strip() if isinstance(prod_name, str) else ""
                    # same-row customer?
                    if isinstance(cust_name, str) and cust_name.strip():
                        q = pd.to_numeric(qty_cell, errors="coerce")
                        if pd.notna(q) and q != 0:
                            recs.append({
                                "Date": period_date,
                                "CustomerCode_ext": str(cust_code).strip() if cust_code is not None else "",
                                "CustomerName": cust_name.strip(),
                                "ProductCode": current_prod_code,
                                "ProductName": current_prod_name,
                                "Quantity": int(q),
                            })
                    continue

                # Detail row under current product
                if current_prod_code and isinstance(cust_name, str) and cust_name.strip():
                    q = pd.to_numeric(qty_cell, errors="coerce")
                    if pd.notna(q) and q != 0:
                        recs.append({
                            "Date": period_date,
                            "CustomerCode_ext": str(cust_code).strip() if cust_code is not None else "",
                            "CustomerName": cust_name.strip(),
                            "ProductCode": current_prod_code,
                            "ProductName": current_prod_name,
                            "Quantity": int(q),
                        })

            return pd.DataFrame(recs)

        frames, parse_log = [], []
        for s in sheets:
            try:
                d = extract_sheet(s)
                n = 0 if d is None else len(d)
                if n:
                    frames.append(d)
                parse_log.append(f"{s}: {n} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not frames:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        df_all = pd.concat(frames, ignore_index=True)

        # -------- 2) Mappings (unique-only; prefer filtered 30010154, then global) --------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str, engine=map_eng)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping",    dtype=str, engine=map_eng)

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30010154"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30010154"].copy()

        cust_f_map = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        cust_g_map = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        sku_f_map  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)
        sku_g_map  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)

        df_all["CustomerCode_norm"] = df_all["CustomerCode_ext"].map(norm_code)
        df_all["CustomerCode"] = df_all["CustomerCode_norm"].map(cust_f_map).fillna(
                                  df_all["CustomerCode_norm"].map(cust_g_map)).fillna("")  # leave blank if unmapped

        df_all["ProductCode_norm"] = df_all["ProductCode"].map(norm_sku)
        df_all["PRT_Product_Code"] = df_all["ProductCode_norm"].map(sku_f_map).fillna(
                                      df_all["ProductCode_norm"].map(sku_g_map)).fillna("")

        # -------- 3) Assemble final + aggregate duplicates --------
        final_fixed = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30010154",
            "GroupName": "亨玖",
            "CustomerCode": df_all["CustomerCode"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
        })

        final_fixed = final_fixed.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName"],
            as_index=False
        )["Quantity"].sum()

        # -------- UI --------
        st.write("✅ Processed Data Preview (first 25 rows):")
        st.dataframe(final_fixed.head(25))

        with st.expander("🔎 Parse summary (per sheet)"):
            st.code("\n".join(parse_log))

        # Download (no headers, no index)
        out_name = "30010154_亨玖_transformation.xlsx"
        final_fixed.to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30010185 瑞星翰德(夜點)":
    import re
    import pandas as pd
    import streamlit as st

    # ---- Uploaders: allow both .xls and .xlsx
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls","xlsx"], key="ruixing_night_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls","xlsx"], key="ruixing_night_map")

    if raw_data_file is not None and mapping_file is not None:
        # -------- Engines (.xls needs xlrd) --------
        def pick_engine(uploaded):
            return "xlrd" if uploaded and uploaded.name.lower().endswith(".xls") else None
        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)

        # -------- Helpers --------
        def parse_period_end(df: pd.DataFrame) -> str | None:
            """
            Find '日期區間:114/07/01~114/07/30' and return end date as YYYYMMDD.
            """
            for r in range(min(30, len(df))):
                for c in range(min(8, df.shape[1])):
                    val = df.iat[r, c]
                    if pd.isna(val):
                        continue
                    s = str(val).strip()
                    m = re.search(r'日期區間[:：]\s*(\d{3})/(\d{2})/(\d{2})\s*~\s*(\d{3})/(\d{2})/(\d{2})', s)
                    if m:
                        y2, m2, d2 = int(m.group(4)) + 1911, int(m.group(5)), int(m.group(6))
                        return f"{y2:04d}{m2:02d}{d2:02d}"
            return None

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # -------- 1) Parse all sheets --------
        xls = pd.ExcelFile(raw_data_file, engine=raw_eng)
        sheets = xls.sheet_names

        def extract_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            date_end = parse_period_end(df)

            # Expect header row: 產品編號 | 品名規格 | 客戶編號 | 客戶名稱 | 數量(...)
            header_row = None
            for r in range(len(df)):
                row = [str(df.iat[r, c]).strip() if (c < df.shape[1] and pd.notna(df.iat[r, c])) else "" for c in range(df.shape[1])]
                if len(row) >= 5 and row[0] == "產品編號" and row[1] == "品名規格" and row[2] == "客戶編號" and row[3] == "客戶名稱" and "數量" in row[4]:
                    header_row = r
                    break
            if header_row is None:
                return pd.DataFrame()

            recs = []
            for r in range(header_row + 1, len(df)):
                prod_code = df.iat[r, 0] if df.shape[1] > 0 else None
                prod_name = df.iat[r, 1] if df.shape[1] > 1 else None
                cust_code = df.iat[r, 2] if df.shape[1] > 2 else None
                cust_name = df.iat[r, 3] if df.shape[1] > 3 else None
                qty       = df.iat[r, 4] if df.shape[1] > 4 else None

                # stop at totals
                if isinstance(prod_code, str) and prod_code.strip().startswith("總計"):
                    break

                q = pd.to_numeric(qty, errors="coerce")
                if pd.isna(q) or q == 0:
                    continue
                if not isinstance(cust_name, str) or not cust_name.strip():
                    continue

                recs.append({
                    "Date": date_end,
                    "CustomerCode_ext": str(cust_code).strip() if cust_code is not None else "",
                    "CustomerName": cust_name.strip(),
                    "ProductCode": str(prod_code).strip().upper() if isinstance(prod_code, str) else "",
                    "ProductName": str(prod_name).strip() if isinstance(prod_name, str) else "",
                    "Quantity": int(q),
                    "Sheet": sheet_name
                })
            return pd.DataFrame(recs)

        frames, parse_log = [], []
        for s in sheets:
            try:
                d = extract_sheet(s)
                n = 0 if d is None else len(d)
                if n:
                    frames.append(d)
                parse_log.append(f"{s}: {n} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not frames:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        df_all = pd.concat(frames, ignore_index=True)

        # -------- 2) Mappings (unique-only; prefer filtered 30010185, then global) --------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str, engine=map_eng)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping",    dtype=str, engine=map_eng)

        cust_map["ASI_CRM_Mapping_Cust_No__c"]    = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"]  = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30010185"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30010185"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)

        df_all["CustomerCode_norm"] = df_all["CustomerCode_ext"].map(norm_code)
        df_all["CustomerCode"] = df_all["CustomerCode_norm"].map(m_cust_f).fillna(
                                  df_all["CustomerCode_norm"].map(m_cust_g)).fillna("")  # leave blank if unmapped

        df_all["ProductCode_norm"] = df_all["ProductCode"].map(norm_sku)
        df_all["PRT_Product_Code"] = df_all["ProductCode_norm"].map(m_sku_f).fillna(
                                      df_all["ProductCode_norm"].map(m_sku_g)).fillna("")

        # -------- 3) Assemble final + aggregate duplicates --------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30010185",
            "GroupName": "瑞星翰德(夜點)",
            "CustomerCode": df_all["CustomerCode"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
            "Sheet": df_all["Sheet"]
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName","Sheet"],
            as_index=False
        )["Quantity"].sum().sort_values(["Sheet","ProductCode","CustomerName"]).reset_index(drop=True)

        # -------- UI --------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["CustomerCode"] == "").sum())
            unmapped_sku  = int((final["PRT_Product_Code"] == "").sum())
            st.code("\n".join(parse_log))
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # -------- Download (no headers, no index; exclude 'Sheet' like usual) --------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "CustomerCode","CustomerName","Date",
                       "PRT_Product_Code","ProductCode","ProductName","Quantity"]
        out_name = "30010185_瑞星翰德_夜點_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30010316 大倉捷":
    import re
    import pandas as pd
    import streamlit as st

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls","xlsx"], key="dakangjie_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls","xlsx"], key="dakangjie_map")

    if raw_data_file is not None and mapping_file is not None:
        # -------- Engines (.xls needs xlrd) --------
        def pick_engine(uploaded):
            return "xlrd" if uploaded and uploaded.name.lower().endswith(".xls") else None
        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)

        # -------- Helpers --------
        def extract_end_date(df: pd.DataFrame) -> str | None:
            """
            Find '貨單日期: 114/07/28 至 114/07/31' and return the right-hand date as YYYYMMDD.
            """
            for r in range(min(12, len(df))):
                for c in range(min(6, df.shape[1])):
                    s = str(df.iat[r, c]) if pd.notna(df.iat[r, c]) else ""
                    m = re.search(r'貨單日期[:：]\s*(\d{3})/(\d{2})/(\d{2})\s*至\s*(\d{3})/(\d{2})/(\d{2})', s)
                    if m:
                        y = int(m.group(4)) + 1911
                        mth = int(m.group(5))
                        d = int(m.group(6))
                        return f"{y:04d}{mth:02d}{d:02d}"
            return None

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # -------- 1) Parse relevant sheets (the report is on 工作表2 in your sample) --------
        xls = pd.ExcelFile(raw_data_file, engine=raw_eng)
        sheets = xls.sheet_names

        def parse_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            date_ymd = extract_end_date(df)
            if date_ymd is None:
                # Skip sheets with no date range (likely not the report)
                return pd.DataFrame()

            recs = []
            current_code = ""
            current_name = ""

            for r in range(len(df)):
                c0 = df.iat[r, 0] if df.shape[1] > 0 else None
                c1 = df.iat[r, 1] if df.shape[1] > 1 else None
                c2 = df.iat[r, 2] if df.shape[1] > 2 else None  # 銷貨數量
                c3 = df.iat[r, 3] if df.shape[1] > 3 else None  # 退貨數量
                c4 = df.iat[r, 4] if df.shape[1] > 4 else None  # 合計數量

                s0 = str(c0).strip() if pd.notna(c0) else ""

                # New product header line e.g. "貨品編號:B0008  貨品名稱:馬爹利(名仕) MARTELL"
                if s0.startswith("貨品編號:"):
                    m = re.search(r'貨品編號[:：]\s*([A-Za-z0-9\-]+)\s+貨品名稱[:：]\s*(.+)', s0)
                    if m:
                        current_code = m.group(1).strip().upper()
                        current_name = m.group(2).strip()
                    else:
                        # Fallback if formatting is odd
                        parts = re.split(r'\s{2,}', s0.replace("貨品編號:", "").replace("貨品名稱:", " ").strip(), maxsplit=1)
                        current_code = parts[0].strip().upper() if parts else ""
                        current_name = parts[1].strip() if len(parts) > 1 else (str(c1).strip() if isinstance(c1, str) else "")
                    continue

                if s0 in ("小計", "客戶編號", "總計"):
                    continue

                # Customer detail line under the current product
                if current_code and s0 and re.match(r'^[A-Za-z0-9\-]+$', s0) and isinstance(c1, str):
                    # Prefer '合計數量' if present; otherwise compute 銷貨 - 退貨
                    qty = pd.to_numeric(c4, errors="coerce")
                    if pd.isna(qty):
                        sales = pd.to_numeric(c2, errors="coerce")
                        returns = pd.to_numeric(c3, errors="coerce")
                        sales = 0 if pd.isna(sales) else sales
                        returns = 0 if pd.isna(returns) else returns
                        qty = sales - returns
                    if pd.notna(qty) and qty != 0:
                        recs.append({
                            "Date": date_ymd,
                            "CustomerCode_ext": s0,
                            "CustomerName": str(c1).strip(),
                            "ProductCode": current_code,
                            "ProductName": current_name,
                            "Quantity": int(qty)
                        })

            return pd.DataFrame(recs)

        frames, parse_log = [], []
        for s in sheets:
            try:
                t = parse_sheet(s)
                if not t.empty:
                    frames.append(t)
                parse_log.append(f"{s}: {len(t)} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not frames:
            st.error("No valid rows found.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        df_all = pd.concat(frames, ignore_index=True)

        # -------- 2) Mappings (unique-only; prefer 30010316, then global) --------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str, engine=map_eng)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping",    dtype=str, engine=map_eng)

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30010316"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30010316"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)

        df_all["CustomerCode_norm"] = df_all["CustomerCode_ext"].map(norm_code)
        df_all["CustomerCode"] = df_all["CustomerCode_norm"].map(m_cust_f).fillna(
                                  df_all["CustomerCode_norm"].map(m_cust_g)).fillna("")  # leave blank if unmapped

        df_all["ProductCode_norm"] = df_all["ProductCode"].map(norm_sku)
        df_all["PRT_Product_Code"] = df_all["ProductCode_norm"].map(m_sku_f).fillna(
                                      df_all["ProductCode_norm"].map(m_sku_g)).fillna("")

        # -------- 3) Assemble final + aggregate duplicates --------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30010316",
            "GroupName": "大倉捷",
            "CustomerCode": df_all["CustomerCode"],
            "CustomerName": df_all["CustomerName"],
            "Date": df_all["Date"],
            "PRT_Product_Code": df_all["PRT_Product_Code"],
            "ProductCode": df_all["ProductCode_norm"],
            "ProductName": df_all["ProductName"],
            "Quantity": df_all["Quantity"].astype(int),
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName"],
            as_index=False
        )["Quantity"].sum().sort_values(["ProductCode","CustomerName"]).reset_index(drop=True)

        # -------- UI --------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["CustomerCode"] == "").sum())
            unmapped_sku  = int((final["PRT_Product_Code"] == "").sum())
            st.code("\n".join(parse_log))
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # -------- Download (no headers, no index) --------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "CustomerCode","CustomerName","Date",
                       "PRT_Product_Code","ProductCode","ProductName","Quantity"]
        out_name = "30010316_大倉捷_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30020076 酒國英豪":
    import re
    import pandas as pd
    import streamlit as st

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls","xlsx"], key="jiuguo_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls","xlsx"], key="jiuguo_map")

    if raw_data_file is not None and mapping_file is not None:
        # -------- Engines (.xls needs xlrd) --------
        def pick_engine(uploaded):
            return "xlrd" if uploaded and uploaded.name.lower().endswith(".xls") else None
        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)

        # -------- Helpers --------
        def is_table_header(df, r: int) -> bool:
            """
            Expect a block header like:
            單據日期 | 單據編號 | 客戶編號 | 客戶簡稱 | 數量(或 數量/Units)
            """
            get = lambda c: str(df.iat[r, c]).strip() if (c < df.shape[1] and pd.notna(df.iat[r, c])) else ""
            c0, c1, c2, c3, c4 = get(0), get(1), get(2), get(3), get(4)
            if c0 == "單據日期" and c1 == "單據編號" and c2 == "客戶編號" and c3 in ("客戶簡稱", "客戶名稱") and ("數量" in c4):
                return True
            return False

        def to_ymd(x) -> str:
            """Convert 'YYYY/MM/DD' or Timestamp to 'YYYYMMDD'. Otherwise blank."""
            if pd.isna(x):
                return ""
            if isinstance(x, pd.Timestamp):
                return x.strftime("%Y%m%d")
            s = str(x).strip()
            m = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$', s)
            if m:
                return f"{int(m.group(1)):04d}{int(m.group(2)):02d}{int(m.group(3)):02d}"
            # last-resort parse
            try:
                dt = pd.to_datetime(s, errors="coerce")
                return dt.strftime("%Y%m%d") if pd.notna(dt) else ""
            except Exception:
                return ""

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            """Only keep mappings where one input maps to exactly one output."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # -------- 1) Parse all sheets (blocks per '貨品編號:' then detail table) --------
        xls = pd.ExcelFile(raw_data_file, engine=raw_eng)
        sheets = xls.sheet_names

        def parse_sheet(sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(raw_data_file, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            recs = []
            current_prod_code, current_prod_name = "", ""

            for r in range(len(df)):
                c0 = df.iat[r, 0] if 0 < df.shape[1] else None
                s0 = str(c0).strip() if pd.notna(c0) else ""

                # Product header e.g. "貨品編號:B07002-004 格蘭利威13年雪莉桶-0.7L"
                if s0.startswith("貨品編號:"):
                    m = re.match(r'貨品編號[:：]\s*([A-Za-z0-9\-]+)\s+(.+)', s0)
                    if m:
                        current_prod_code = m.group(1).strip().upper()
                        current_prod_name = m.group(2).strip()
                    else:
                        current_prod_code = s0.split("貨品編號:")[-1].strip().upper()
                        current_prod_name = ""
                    continue

                # Detail table header for this product block
                if is_table_header(df, r):
                    i = r + 1
                    while i < len(df):
                        row = df.iloc[i]
                        # stop conditions: blank row, a new report title, or next product header
                        s00 = str(row[0]).strip() if pd.notna(row[0]) else ""
                        if (pd.isna(row[0]) and pd.isna(row[1]) and pd.isna(row[2]) and pd.isna(row[3])) \
                           or s00.startswith("酒國英豪洋酒有限公司") or s00.startswith("貨品編號:"):
                            break

                        date_cell = row[0]
                        doc_no    = str(row[1]).strip() if pd.notna(row[1]) else ""
                        cust_ext  = str(row[2]).strip() if pd.notna(row[2]) else ""
                        cust_name = str(row[3]).strip() if pd.notna(row[3]) else ""
                        qty_val   = pd.to_numeric(row[4], errors="coerce")

                        if current_prod_code and cust_name and pd.notna(qty_val) and qty_val != 0:
                            ymd = to_ymd(date_cell)

                            # Return logic: if doc number indicates return, force negative
                            doc_mark = doc_no or ""
                            is_return = any(x in doc_mark for x in ("銷退", "退回", "退貨", "銷售退回"))
                            qty = -abs(int(qty_val)) if is_return else int(qty_val)

                            recs.append({
                                "Date": ymd,
                                "CustomerCode_ext": cust_ext,
                                "CustomerName": cust_name,
                                "ProductCode": current_prod_code,
                                "ProductName": current_prod_name,
                                "Quantity": qty,
                                "DocNo": doc_no,
                                "Sheet": sheet_name
                            })
                        i += 1

            return pd.DataFrame(recs)

        frames, parse_log = [], []
        for s in sheets:
            try:
                part = parse_sheet(s)
                if not part.empty:
                    frames.append(part)
                parse_log.append(f"{s}: {len(part)} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not frames:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        raw_extracted = pd.concat(frames, ignore_index=True)

        # -------- 2) Mappings (unique-only; prefer 30020076, then global) --------
        cust_map = pd.read_excel(mapping_file, sheet_name="Customer Mapping", dtype=str, engine=map_eng)
        sku_map  = pd.read_excel(mapping_file, sheet_name="SKU Mapping",    dtype=str, engine=map_eng)

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30020076"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30020076"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",           norm_sku)

        df_m = raw_extracted.copy()
        df_m["CustomerCode_norm"] = df_m["CustomerCode_ext"].map(norm_code)
        df_m["CustomerCode"] = df_m["CustomerCode_norm"].map(m_cust_f).fillna(
                                df_m["CustomerCode_norm"].map(m_cust_g)).fillna("")   # leave blank if unmapped

        df_m["ProductCode_norm"] = df_m["ProductCode"].map(norm_sku)
        df_m["PRT_Product_Code"] = df_m["ProductCode_norm"].map(m_sku_f).fillna(
                                    df_m["ProductCode_norm"].map(m_sku_g)).fillna("")

        # -------- 3) Assemble final + aggregate (keep DocNo in group to avoid merging invoices) --------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30020076",
            "GroupName": "酒國英豪",
            "CustomerCode": df_m["CustomerCode"],
            "CustomerName": df_m["CustomerName"],
            "Date": df_m["Date"],
            "PRT_Product_Code": df_m["PRT_Product_Code"],
            "ProductCode": df_m["ProductCode_norm"],
            "ProductName": df_m["ProductName"],
            "Quantity": df_m["Quantity"].astype(int),
            "DocNo": df_m["DocNo"],
            "Sheet": df_m["Sheet"]
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName","DocNo","Sheet"],
            as_index=False
        )["Quantity"].sum().sort_values(["Date","ProductCode","CustomerName","DocNo"]).reset_index(drop=True)

        # -------- UI --------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["CustomerCode"] == "").sum())
            unmapped_sku  = int((final["PRT_Product_Code"] == "").sum())
            st.code("\n".join(parse_log))
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # -------- Download (no headers, no index). Exclude DocNo/Sheet per your standard. --------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "CustomerCode","CustomerName","Date",
                       "PRT_Product_Code","ProductCode","ProductName","Quantity"]
        out_name = "30020076_酒國英豪_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30030021 合歡 ON":
    import re
    import pandas as pd
    import streamlit as st
    from io import BytesIO

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls", "xlsx"], key="hehuan_on_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls", "xlsx"], key="hehuan_on_map")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------- Engine helpers (fixes xlrd reading .xlsx error) ----------------
        def pick_engine(uploaded):
            name = (uploaded.name or "").lower()
            if name.endswith(".xlsx") or ".xlsx" in name:
                return "openpyxl"   # force openpyxl for any .xlsx (even weird names like .xls.xlsx)
            if name.endswith(".xls"):
                return "xlrd"
            return None  # let pandas choose

        def to_buffer(uploaded_file):
            data = uploaded_file.read()
            return data, BytesIO(data)  # (bytes, reusable buffer)

        def read_excel_safe(buf_bytes, sheet_name=0, header=None, engine=None, **kwargs):
            # try chosen engine
            try:
                return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=engine, **kwargs)
            except Exception:
                # try without engine (let pandas decide)
                try:
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, **kwargs)
                except Exception:
                    # last fallback: alternate engine
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=alt, **kwargs)

        def excel_file_safe(buf_bytes, engine=None):
            try:
                return pd.ExcelFile(BytesIO(buf_bytes), engine=engine)
            except Exception:
                try:
                    return pd.ExcelFile(BytesIO(buf_bytes))  # let pandas decide
                except Exception:
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.ExcelFile(BytesIO(buf_bytes), engine=alt)

        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)
        raw_bytes, _ = to_buffer(raw_data_file)
        map_bytes, _  = to_buffer(mapping_file)

        # ---------------- Utilities ----------------
        def minguo_to_ymd(s: object) -> str:
            """Convert Minguo date '114/05/07' → '20250507'. Return '' if not match."""
            if pd.isna(s):
                return ""
            s = str(s).strip()
            m = re.match(r"^(\d{3})/(\d{2})/(\d{2})$", s)
            if not m:
                return ""
            y = int(m.group(1)) + 1911
            return f"{y:04d}{int(m.group(2)):02d}{int(m.group(3)):02d}"

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            """Only keep mappings where one input maps to exactly one output."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # ---------------- 1) Parse all sheets ----------------
        xls = excel_file_safe(raw_bytes, engine=raw_eng)
        sheets = xls.sheet_names

        def parse_sheet(sheet_name: str) -> pd.DataFrame:
            df = read_excel_safe(raw_bytes, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            recs = []
            current_prod_code = ""
            current_prod_name = ""

            for r in range(len(df)):
                c0 = df.iat[r, 0] if 0 < df.shape[1] else None
                s0 = str(c0).strip() if pd.notna(c0) else ""

                # Product header line: "產品編號：" in col0, code in col1, name in col3
                if s0.startswith("產品編號"):
                    code = str(df.iat[r, 1]).strip().upper() if (df.shape[1] > 1 and pd.notna(df.iat[r, 1])) else ""
                    name = str(df.iat[r, 3]).strip() if (df.shape[1] > 3 and pd.notna(df.iat[r, 3])) else ""
                    name = name.lstrip("[").strip()
                    current_prod_code, current_prod_name = code, name
                    continue

                # Detail line begins with Minguo date like "114/05/07"
                if re.match(r"^\d{3}/\d{2}/\d{2}$", s0):
                    date_ymd = minguo_to_ymd(s0)
                    doc_no   = str(df.iat[r, 1]).strip() if (df.shape[1] > 1 and pd.notna(df.iat[r, 1])) else ""
                    cust_ext = str(df.iat[r, 2]).strip() if (df.shape[1] > 2 and pd.notna(df.iat[r, 2])) else ""
                    cust_nm  = str(df.iat[r, 3]).strip() if (df.shape[1] > 3 and pd.notna(df.iat[r, 3])) else ""
                    qty      = pd.to_numeric(df.iat[r, 5] if df.shape[1] > 5 else None, errors="coerce")

                    if current_prod_code and cust_nm and pd.notna(qty) and qty != 0:
                        recs.append({
                            "Date": date_ymd,
                            "CustomerCode_ext": cust_ext,
                            "CustomerName": cust_nm,
                            "ProductCode": current_prod_code,
                            "ProductName": current_prod_name,
                            "Quantity": int(qty),
                            "DocNo": doc_no,
                            "Sheet": sheet_name
                        })

            return pd.DataFrame(recs)

        parts, parse_log = [], []
        for s in sheets:
            try:
                p = parse_sheet(s)
                if not p.empty:
                    parts.append(p)
                parse_log.append(f"{s}: {len(p)} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not parts:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        raw_extracted = pd.concat(parts, ignore_index=True)

        # ---------------- 2) Mappings (unique-only; prefer 30030021, then global) ----------------
        # IMPORTANT: read with header=0 so column names exist
        cust_map = read_excel_safe(map_bytes, sheet_name="Customer Mapping", dtype=str, header=0, engine=map_eng)
        sku_map  = read_excel_safe(map_bytes, sheet_name="SKU Mapping",    dtype=str, header=0, engine=map_eng)

        # normalize column labels (trim stray whitespace)
        cust_map.columns = cust_map.columns.map(lambda x: str(x).strip())
        sku_map.columns  = sku_map.columns.map(lambda x: str(x).strip())

        # proceed with expected columns
        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30030021"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30030021"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)

        df_m = raw_extracted.copy()
        df_m["CustomerCode_norm"] = df_m["CustomerCode_ext"].map(norm_code)
        df_m["CustomerCode"] = df_m["CustomerCode_norm"].map(m_cust_f).fillna(
                                df_m["CustomerCode_norm"].map(m_cust_g)
                              ).fillna("")  # leave blank if unmapped (non-forced)

        df_m["ProductCode_norm"] = df_m["ProductCode"].map(norm_sku)
        df_m["PRT_Product_Code"] = df_m["ProductCode_norm"].map(m_sku_f).fillna(
                                    df_m["ProductCode_norm"].map(m_sku_g)
                                  ).fillna("")

        # ---------------- 3) Assemble final + aggregate duplicates ----------------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030021",
            "GroupName": "合歡 ON",
            "CustomerCode": df_m["CustomerCode"],
            "CustomerName": df_m["CustomerName"],
            "Date": df_m["Date"],
            "PRT_Product_Code": df_m["PRT_Product_Code"],
            "ProductCode": df_m["ProductCode_norm"],
            "ProductName": df_m["ProductName"],
            "Quantity": df_m["Quantity"].astype(int),
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName"],
            as_index=False
        )["Quantity"].sum().sort_values(["Date","ProductCode","CustomerName"]).reset_index(drop=True)

        # ---------------- UI ----------------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["CustomerCode"] == "").sum())
            unmapped_sku  = int((final["PRT_Product_Code"] == "").sum())
            st.code("\n".join(parse_log))
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # ---------------- Download (no headers, no index) ----------------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "CustomerCode","CustomerName","Date",
                       "PRT_Product_Code","ProductCode","ProductName","Quantity"]
        out_name = "30030021_合歡ON_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30030083 東瀛":
    import re
    import pandas as pd
    import streamlit as st
    from io import BytesIO

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls", "xlsx"], key="dongying_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls", "xlsx"], key="dongying_map")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------- Engine helpers (robust for xls/xlsx & weird filenames) ----------------
        def pick_engine(uploaded):
            name = (uploaded.name or "").lower()
            if name.endswith(".xlsx") or ".xlsx" in name:
                return "openpyxl"   # force openpyxl for any .xlsx (even names like .xls.xlsx)
            if name.endswith(".xls"):
                return "xlrd"
            return None  # let pandas choose

        def to_buffer(uploaded_file):
            data = uploaded_file.read()
            return data, BytesIO(data)  # (bytes, reusable buffer)

        def read_excel_safe(buf_bytes, sheet_name=0, header=None, engine=None, **kwargs):
            # try chosen engine
            try:
                return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=engine, **kwargs)
            except Exception:
                # try without engine (let pandas decide)
                try:
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, **kwargs)
                except Exception:
                    # last fallback: alternate engine
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=alt, **kwargs)

        def excel_file_safe(buf_bytes, engine=None):
            try:
                return pd.ExcelFile(BytesIO(buf_bytes), engine=engine)
            except Exception:
                try:
                    return pd.ExcelFile(BytesIO(buf_bytes))  # let pandas decide
                except Exception:
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.ExcelFile(BytesIO(buf_bytes), engine=alt)

        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)
        raw_bytes, _ = to_buffer(raw_data_file)
        map_bytes, _  = to_buffer(mapping_file)

        # ---------------- Utilities ----------------
        def minguo_to_ymd(s: object) -> str:
            """Convert Minguo date '114/07/24' → '20250724'. Return '' if not match."""
            if pd.isna(s):
                return ""
            s = str(s).strip()
            m = re.match(r"^(\d{3})/(\d{2})/(\d{2})$", s)
            if not m:
                return ""
            y = int(m.group(1)) + 1911
            return f"{y:04d}{int(m.group(2)):02d}{int(m.group(3)):02d}"

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            """Only keep mappings where one input maps to exactly one output."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # ---------------- 1) Parse all sheets ----------------
        xls = excel_file_safe(raw_bytes, engine=raw_eng)
        sheets = xls.sheet_names

        def parse_sheet(sheet_name: str) -> pd.DataFrame:
            df = read_excel_safe(raw_bytes, sheet_name=sheet_name, header=None, engine=raw_eng)
            if df.empty:
                return pd.DataFrame()

            recs = []
            current_code = ""
            current_name = ""

            for r in range(len(df)):
                s0 = str(df.iat[r, 0]).strip() if (df.shape[1] > 0 and pd.notna(df.iat[r, 0])) else ""

                # Product header: e.g. "貨品編號:D0530-0  貨品名稱:格蘭利威12年-盒裝"
                if s0.startswith("貨品編號"):
                    m = re.search(r"貨品編號[:：]\s*([A-Za-z0-9\-]+)\s+(?:貨品名稱[:：])?\s*(.+)", s0)
                    if m:
                        current_code = m.group(1).strip().upper()
                        current_name = m.group(2).strip()
                    else:
                        # fallback if only code present on line
                        current_code = s0.split("貨品編號")[-1].lstrip(":：").strip().upper()
                        current_name = ""
                    continue

                # Detail rows: col2 is Minguo date "114/07/24"
                if df.shape[1] > 4:
                    date_cell = df.iat[r, 2]
                    date_str = str(date_cell).strip() if pd.notna(date_cell) else ""
                    if re.match(r"^\d{3}/\d{2}/\d{2}$", date_str):
                        cust_code = str(df.iat[r, 0]).strip() if pd.notna(df.iat[r, 0]) else ""
                        cust_name = str(df.iat[r, 1]).strip() if pd.notna(df.iat[r, 1]) else ""
                        # skip headers/subtotals
                        if cust_code in ("小計", "總計", "客戶編號", "") or cust_name in ("客戶名稱", ""):
                            continue

                        doc_no = str(df.iat[r, 3]).strip() if pd.notna(df.iat[r, 3]) else ""
                        qty = pd.to_numeric(df.iat[r, 4], errors="coerce")

                        if current_code and cust_name and pd.notna(qty) and qty != 0:
                            ymd = minguo_to_ymd(date_str)
                            is_return = any(mark in doc_no for mark in ("退", "銷退", "退回", "退貨"))
                            q = -abs(int(qty)) if is_return else int(qty)

                            recs.append({
                                "Date": ymd,
                                "CustomerCode_ext": cust_code,
                                "CustomerName": cust_name,
                                "ProductCode": current_code,
                                "ProductName": current_name,
                                "Quantity": q,
                                "DocNo": doc_no,
                                "Sheet": sheet_name
                            })
            return pd.DataFrame(recs)

        parts, parse_log = [], []
        for s in sheets:
            try:
                p = parse_sheet(s)
                if not p.empty:
                    parts.append(p)
                parse_log.append(f"{s}: {len(p)} rows")
            except Exception as e:
                parse_log.append(f"{s}: ERROR → {e}")

        if not parts:
            st.error("No valid rows found in any sheet.\n\nParse summary:\n" + "\n".join(parse_log))
            st.stop()

        raw_extracted = pd.concat(parts, ignore_index=True)

        # ---------------- 2) Mappings (unique-only; prefer 30030083, then global) ----------------
        cust_map = read_excel_safe(map_bytes, sheet_name="Customer Mapping", dtype=str, header=0, engine=map_eng)
        sku_map  = read_excel_safe(map_bytes, sheet_name="SKU Mapping",    dtype=str, header=0, engine=map_eng)

        # trim column labels
        cust_map.columns = cust_map.columns.map(lambda x: str(x).strip())
        sku_map.columns  = sku_map.columns.map(lambda x: str(x).strip())

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30030083"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30030083"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)

        df_m = raw_extracted.copy()
        df_m["CustomerCode_norm"] = df_m["CustomerCode_ext"].map(norm_code)
        df_m["CustomerCode"] = df_m["CustomerCode_norm"].map(m_cust_f).fillna(
                                df_m["CustomerCode_norm"].map(m_cust_g)
                              ).fillna("")  # leave blank if unmapped (non-forced)

        df_m["ProductCode_norm"] = df_m["ProductCode"].map(norm_sku)
        df_m["PRT_Product_Code"] = df_m["ProductCode_norm"].map(m_sku_f).fillna(
                                    df_m["ProductCode_norm"].map(m_sku_g)
                                  ).fillna("")

        # ---------------- 3) Assemble final + aggregate (keep DocNo to avoid merging invoices) ----------------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030083",
            "GroupName": "東瀛",
            "CustomerCode": df_m["CustomerCode"],
            "CustomerName": df_m["CustomerName"],
            "Date": df_m["Date"],
            "PRT_Product_Code": df_m["PRT_Product_Code"],
            "ProductCode": df_m["ProductCode_norm"],
            "ProductName": df_m["ProductName"],
            "Quantity": df_m["Quantity"].astype(int),
            "DocNo": df_m["DocNo"],
            "Sheet": df_m["Sheet"]
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "CustomerCode","CustomerName","Date",
             "PRT_Product_Code","ProductCode","ProductName","DocNo","Sheet"],
            as_index=False
        )["Quantity"].sum().sort_values(["Date","ProductCode","CustomerName","DocNo"]).reset_index(drop=True)

        # ---------------- UI ----------------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["CustomerCode"] == "").sum())
            unmapped_sku  = int((final["PRT_Product_Code"] == "").sum())
            st.code("\n".join(parse_log))
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # ---------------- Download (no headers, no index). Exclude DocNo/Sheet per your standard. ----------------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "CustomerCode","CustomerName","Date",
                       "PRT_Product_Code","ProductCode","ProductName","Quantity"]
        out_name = "30030083_東瀛_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30030084 華恩":
    import re
    import pandas as pd
    import streamlit as st
    from io import BytesIO

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls", "xlsx"], key="huaen_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls", "xlsx"], key="huaen_map")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------- Engine helpers (robust for xls/xlsx & weird filenames) ----------------
        def pick_engine(uploaded):
            name = (uploaded.name or "").lower()
            if name.endswith(".xlsx") or ".xlsx" in name:
                return "openpyxl"   # force openpyxl for any .xlsx (even names like .xls.xlsx)
            if name.endswith(".xls"):
                return "xlrd"
            return None  # let pandas choose

        def to_buffer(uploaded_file):
            data = uploaded_file.read()
            return data, BytesIO(data)  # (bytes, reusable buffer)

        def read_excel_safe(buf_bytes, sheet_name=0, header=None, engine=None, **kwargs):
            # try chosen engine
            try:
                return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=engine, **kwargs)
            except Exception:
                # try without engine (let pandas decide)
                try:
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, **kwargs)
                except Exception:
                    # last fallback: alternate engine
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=alt, **kwargs)

        def excel_file_safe(buf_bytes, engine=None):
            try:
                return pd.ExcelFile(BytesIO(buf_bytes), engine=engine)
            except Exception:
                try:
                    return pd.ExcelFile(BytesIO(buf_bytes))  # let pandas decide
                except Exception:
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.ExcelFile(BytesIO(buf_bytes), engine=alt)

        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)
        raw_bytes, _ = to_buffer(raw_data_file)
        map_bytes, _  = to_buffer(mapping_file)

        # ---------------- Helpers ----------------
        def parse_date_range(s):
            # "統計日期：2025/06/23至2025/06/30" -> end date "20250630"
            s = "" if s is None else str(s)
            m = re.search(r'(\d{4})/(\d{2})/(\d{2})\s*至\s*(\d{4})/(\d{2})/(\d{2})', s)
            if m:
                y2, m2, d2 = int(m.group(4)), int(m.group(5)), int(m.group(6))
                return f"{y2:04d}{m2:02d}{d2:02d}"
            return ""

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            """Only keep mappings where one input maps to exactly one output."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # ---------------- 1) Parse raw ("銷售") ----------------
        xls = excel_file_safe(raw_bytes, engine=raw_eng)
        sheet_name = "銷售" if "銷售" in xls.sheet_names else xls.sheet_names[0]
        df = read_excel_safe(raw_bytes, sheet_name=sheet_name, header=None, engine=raw_eng)

        # Pass 1: collect blocks of customers per product (with end date)
        blocks = []  # {"date": ymd, "product_code": code, "customers": [(ext_code, name), ...]}
        i = 0
        while i < len(df):
            cell0 = df.iloc[i, 0] if 0 in df.columns else None
            if "統計日期" in (str(cell0) if pd.notna(cell0) else ""):
                date_ymd = parse_date_range(cell0)
                # find "客戶名稱" row
                j = i + 1
                while j < len(df) and str(df.iloc[j, 0]).strip() != "客戶名稱":
                    j += 1
                if j >= len(df):
                    i += 1
                    continue
                # from j+1 downward: rows with customer code/name + product code in col5
                k = j + 1
                customers = []
                product_code = ""
                while k < len(df):
                    ccode = str(df.iloc[k, 0]).strip() if pd.notna(df.iloc[k, 0]) else ""
                    cname = str(df.iloc[k, 1]).strip() if pd.notna(df.iloc[k, 1]) else ""
                    pcode = str(df.iloc[k, 5]).strip() if (df.shape[1] > 5 and pd.notna(df.iloc[k, 5])) else ""
                    if ccode and cname and pcode and ccode != "客戶名稱":
                        if not product_code:
                            product_code = pcode.upper()
                        customers.append((ccode, cname))
                        k += 1
                        continue
                    # stop on blank separator
                    if (not ccode and not cname):
                        break
                    k += 1
                if product_code and customers:
                    blocks.append({"date": date_ymd, "product_code": product_code, "customers": customers})
                i = k
                continue
            i += 1

        # Pass 2: for each "品名規格 ... 銷量" section, collect (name, qty) rows in order
        qty_sections = []  # list of list[(product_name, qty)]
        i = 0
        while i < len(df):
            row = df.iloc[i]
            found = False
            for c in range(df.shape[1]):
                if str(row[c]).strip() == "品名規格":
                    # quantity often located c+4 columns to the right (observed: name at col6, qty at col10)
                    qty_col = c + 4 if (c + 4) < df.shape[1] else df.shape[1] - 1
                    names_qty = []
                    r = i + 1
                    while r < len(df):
                        name = str(df.iat[r, c]).strip() if pd.notna(df.iat[r, c]) else ""
                        # stop at subtotal/blank
                        if not name:
                            break
                        # next column might contain "合計 :" marker
                        next_cell = str(df.iat[r, c + 1]).strip() if (c + 1) < df.shape[1] and pd.notna(df.iat[r, c + 1]) else ""
                        if "合計" in next_cell:
                            break
                        qty = pd.to_numeric(df.iat[r, qty_col] if (qty_col) < df.shape[1] else None, errors="coerce")
                        if pd.notna(qty):
                            names_qty.append((name, int(qty)))
                        r += 1
                    if names_qty:
                        qty_sections.append(names_qty)
                    i = r
                    found = True
                    break
            if not found:
                i += 1

        # Pair blocks with quantity sections by order; zip customers with names_qty
        records = []
        pair_count = min(len(blocks), len(qty_sections))
        for idx in range(pair_count):
            blk = blocks[idx]
            pairs = qty_sections[idx]
            customers = blk["customers"]
            for (cust_code, cust_name), (pname, qty) in zip(customers, pairs):
                records.append({
                    "Date": blk["date"],
                    "CustomerCode_ext": cust_code,
                    "CustomerName": cust_name,
                    "ProductCode": blk["product_code"],
                    "ProductName": pname,
                    "Quantity": int(qty),
                    "DocNo": "",
                    "Sheet": sheet_name
                })

        raw_extracted = pd.DataFrame(records)
        if raw_extracted.empty:
            st.error("No detail rows parsed. Please verify the layout.")
            st.stop()

        # ---------------- 2) Mappings (unique-only; prefer 30030084, then global; leave blank if unmapped) ----------------
        cust_map = read_excel_safe(map_bytes, sheet_name="Customer Mapping", dtype=str, header=0, engine=map_eng)
        sku_map  = read_excel_safe(map_bytes, sheet_name="SKU Mapping",    dtype=str, header=0, engine=map_eng)
        cust_map.columns = cust_map.columns.map(lambda x: str(x).strip())
        sku_map.columns  = sku_map.columns.map(lambda x: str(x).strip())

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30030084"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30030084"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)

        df_m = raw_extracted.copy()
        df_m["CustomerCode_norm"] = df_m["CustomerCode_ext"].map(norm_code)
        df_m["CustomerCode"] = df_m["CustomerCode_norm"].map(m_cust_f).fillna(
                                df_m["CustomerCode_norm"].map(m_cust_g)
                              ).fillna("")  # leave blank if unmapped (non-forced)
        df_m["ProductCode_norm"] = df_m["ProductCode"].map(norm_sku)
        df_m["PRT_Product_Code"] = df_m["ProductCode_norm"].map(m_sku_f).fillna(
                                    df_m["ProductCode_norm"].map(m_sku_g)
                                  ).fillna("")

        # ---------------- 3) Assemble final + aggregate duplicates ----------------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030084",
            "GroupName": "華恩",
            "Customer Code": df_m["CustomerCode"],
            "Customer Name": df_m["CustomerName"],
            "Date": df_m["Date"],
            "PRT Product Code": df_m["PRT_Product_Code"],
            "Product Code": df_m["ProductCode_norm"],
            "Product Name": df_m["ProductName"],
            "Number of Bottles": df_m["Quantity"].astype(int),
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "Customer Code","Customer Name","Date",
             "PRT Product Code","Product Code","Product Name"],
            as_index=False
        )["Number of Bottles"].sum().sort_values(["Date","Product Code","Customer Name"]).reset_index(drop=True)

        # ---------------- UI ----------------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["Customer Code"] == "").sum())
            unmapped_sku  = int((final["PRT Product Code"] == "").sum())
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")
            st.caption("Customers are paired to quantities in order within each product block (left list ↔ right ‘品名規格/銷量’ section).")

        # ---------------- Download (no headers, no index) ----------------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "Customer Code","Customer Name","Date",
                       "PRT Product Code","Product Code","Product Name","Number of Bottles"]
        out_name = "30030084_華恩_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)

elif transformation_choice == "30030106 明輝":
    import re
    import pandas as pd
    import streamlit as st
    from io import BytesIO

    # ---- Uploaders (allow .xls / .xlsx)
    raw_data_file = st.file_uploader("Upload Raw Sales Data (.xls/.xlsx)", type=["xls", "xlsx"], key="minghui_raw")
    mapping_file  = st.file_uploader("Upload Mapping File (.xls/.xlsx)",  type=["xls", "xlsx"], key="minghui_map")

    if raw_data_file is not None and mapping_file is not None:
        # ---------------- Engine helpers (robust for xls/xlsx & weird filenames) ----------------
        def pick_engine(uploaded):
            name = (uploaded.name or "").lower()
            if name.endswith(".xlsx") or ".xlsx" in name:
                return "openpyxl"   # force openpyxl for any .xlsx (even names like .xls.xlsx)
            if name.endswith(".xls"):
                return "xlrd"
            return None  # let pandas choose

        def to_buffer(uploaded_file):
            data = uploaded_file.read()
            return data, BytesIO(data)  # (bytes, reusable buffer)

        def read_excel_safe(buf_bytes, sheet_name=0, header=None, engine=None, **kwargs):
            # try chosen engine
            try:
                return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=engine, **kwargs)
            except Exception:
                # try without engine (let pandas decide)
                try:
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, **kwargs)
                except Exception:
                    # last fallback: alternate engine
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.read_excel(BytesIO(buf_bytes), sheet_name=sheet_name, header=header, engine=alt, **kwargs)

        def excel_file_safe(buf_bytes, engine=None):
            try:
                return pd.ExcelFile(BytesIO(buf_bytes), engine=engine)
            except Exception:
                try:
                    return pd.ExcelFile(BytesIO(buf_bytes))  # let pandas decide
                except Exception:
                    alt = "openpyxl" if engine == "xlrd" else "xlrd"
                    return pd.ExcelFile(BytesIO(buf_bytes), engine=alt)

        raw_eng = pick_engine(raw_data_file)
        map_eng = pick_engine(mapping_file)
        raw_bytes, _ = to_buffer(raw_data_file)
        map_bytes, _  = to_buffer(mapping_file)

        # ---------------- Utilities ----------------
        def minguo_to_ymd(s: object) -> str:
            """Convert Minguo date '114/07/01' → '20250701'. Return '' if not match."""
            if pd.isna(s):
                return ""
            s = str(s).strip()
            m = re.match(r"^(\d{3})/(\d{2})/(\d{2})$", s)
            if not m:
                return ""
            y = int(m.group(1)) + 1911
            return f"{y:04d}{int(m.group(2)):02d}{int(m.group(3)):02d}"

        def unique_only_map(df, key_col, val_col, normalize=lambda s: s):
            """Only keep mappings where one input maps to exactly one output."""
            tmp = df[[key_col, val_col]].dropna().copy()
            tmp["key"] = tmp[key_col].astype(str).map(normalize)
            tmp["val"] = tmp[val_col].astype(str).str.strip()
            counts = tmp.groupby("key")["val"].nunique().reset_index(name="n")
            uniq = set(counts[counts["n"] == 1]["key"])
            tmp = tmp[tmp["key"].isin(uniq)].drop_duplicates(subset="key", keep="first")
            return dict(zip(tmp["key"], tmp["val"]))

        norm_code = lambda s: str(s).strip().upper().replace(" ", "").replace(".0", "")
        norm_sku  = lambda s: str(s).strip().upper()

        # ---------------- 1) Parse raw (Sheet1) ----------------
        # Sheet1 columns (observed):
        # 0: 日期(民國) | 1: 客戶編號 | 2: 客戶簡稱 | 3: 單別 | 4: 產品編號 | 5: 名稱規格 | 6: 銷售數量 | 7: 贈送數量 | ...
        df = read_excel_safe(raw_bytes, sheet_name="Sheet1", header=None, engine=raw_eng)

        records = []
        for r in range(len(df)):
            date_cell = df.iat[r, 0] if df.shape[1] > 0 else None
            date_str = str(date_cell).strip() if pd.notna(date_cell) else ""
            if re.match(r"^\d{3}/\d{2}/\d{2}$", date_str):
                cust_ext = str(df.iat[r, 1]).strip() if pd.notna(df.iat[r, 1]) else ""
                cust_nm  = str(df.iat[r, 2]).strip() if pd.notna(df.iat[r, 2]) else ""
                prod_cd  = str(df.iat[r, 4]).strip().upper() if (df.shape[1] > 4 and pd.notna(df.iat[r, 4])) else ""
                prod_nm  = str(df.iat[r, 5]).strip() if (df.shape[1] > 5 and pd.notna(df.iat[r, 5])) else ""
                qty      = pd.to_numeric(df.iat[r, 6] if df.shape[1] > 6 else None, errors="coerce")

                if cust_nm and prod_cd and pd.notna(qty) and qty != 0:
                    records.append({
                        "Date": minguo_to_ymd(date_str),
                        "CustomerCode_ext": cust_ext,
                        "CustomerName": cust_nm,
                        "ProductCode": prod_cd,
                        "ProductName": prod_nm,
                        "Quantity": int(qty)
                    })

        raw_extracted = pd.DataFrame(records)
        if raw_extracted.empty:
            st.error("No valid rows parsed from the raw file.")
            st.stop()

        # ---------------- 2) Mappings (unique-only; prefer 30030106, then global; leave blank if unmapped) ----------------
        cust_map = read_excel_safe(map_bytes, sheet_name="Customer Mapping", dtype=str, header=0, engine=map_eng)
        sku_map  = read_excel_safe(map_bytes, sheet_name="SKU Mapping",    dtype=str, header=0, engine=map_eng)

        # trim column labels
        cust_map.columns = cust_map.columns.map(lambda x: str(x).strip())
        sku_map.columns  = sku_map.columns.map(lambda x: str(x).strip())

        cust_map["ASI_CRM_Mapping_Cust_No__c"]   = cust_map["ASI_CRM_Mapping_Cust_No__c"].astype(str).str.replace(r"\.0$", "", regex=True)
        sku_map["ASI_CRM_Mapping_Cust_Code__c"] = sku_map["ASI_CRM_Mapping_Cust_Code__c"].astype(str).str.replace(r"\.0$", "", regex=True)

        cust_f = cust_map[cust_map["ASI_CRM_Mapping_Cust_No__c"] == "30030106"].copy()
        sku_f  = sku_map[sku_map["ASI_CRM_Mapping_Cust_Code__c"] == "30030106"].copy()

        m_cust_f = unique_only_map(cust_f,  "ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_cust_g = unique_only_map(cust_map,"ASI_CRM_Offtake_Customer_No__c", "ASI_CRM_JDE_Cust_No_Formula__c", norm_code)
        m_sku_f  = unique_only_map(sku_f,   "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)
        m_sku_g  = unique_only_map(sku_map, "ASI_CRM_Offtake_Product__c",     "ASI_CRM_SKU_Code__c",            norm_sku)

        df_m = raw_extracted.copy()
        df_m["CustomerCode_norm"] = df_m["CustomerCode_ext"].map(norm_code)
        df_m["CustomerCode"] = df_m["CustomerCode_norm"].map(m_cust_f).fillna(
                                df_m["CustomerCode_norm"].map(m_cust_g)
                              ).fillna("")  # leave blank if unmapped (non-forced)

        df_m["ProductCode_norm"] = df_m["ProductCode"].map(norm_sku)
        df_m["PRT_Product_Code"] = df_m["ProductCode_norm"].map(m_sku_f).fillna(
                                    df_m["ProductCode_norm"].map(m_sku_g)
                                  ).fillna("")

        # ---------------- 3) Assemble final + aggregate duplicates ----------------
        final = pd.DataFrame({
            "Type": "INV",
            "Action": "U",
            "GroupCode": "30030106",
            "GroupName": "明輝",
            "Customer Code": df_m["CustomerCode"],
            "Customer Name": df_m["CustomerName"],
            "Date": df_m["Date"],
            "PRT Product Code": df_m["PRT_Product_Code"],
            "Product Code": df_m["ProductCode_norm"],
            "Product Name": df_m["ProductName"],
            "Number of Bottles": df_m["Quantity"].astype(int),
        })

        final = final.groupby(
            ["Type","Action","GroupCode","GroupName",
             "Customer Code","Customer Name","Date",
             "PRT Product Code","Product Code","Product Name"],
            as_index=False
        )["Number of Bottles"].sum().sort_values(["Date","Product Code","Customer Name"]).reset_index(drop=True)

        # ---------------- UI ----------------
        st.write("✅ Processed Data Preview (first 20 rows):")
        st.dataframe(final.head(20))

        with st.expander("🔎 Parse & Mapping Summary"):
            unmapped_cust = int((final["Customer Code"] == "").sum())
            unmapped_sku  = int((final["PRT Product Code"] == "").sum())
            st.write(f"Total rows: {len(final)} | Unmapped customers: {unmapped_cust} | Unmapped SKUs: {unmapped_sku}")

        # ---------------- Download (no headers, no index) ----------------
        export_cols = ["Type","Action","GroupCode","GroupName",
                       "Customer Code","Customer Name","Date",
                       "PRT Product Code","Product Code","Product Name","Number of Bottles"]
        out_name = "30030106_明輝_transformation.xlsx"
        final[export_cols].to_excel(out_name, index=False, header=False)
        with open(out_name, "rb") as f:
            st.download_button("📥 Download Processed File", f, file_name=out_name)
